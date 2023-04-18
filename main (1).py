import sys
import time
import typing
from turtle import color

import PyQt5
from PyQt5 import uic
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtGui import QBrush
from PyQt5.QtGui import QFont, QPalette, QColor
from PyQt5.QtGui import QIcon
from PyQt5.QtGui import QStandardItem
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtWidgets import *
from itertools import product
from PyQt5 import Qt
from PyQt5 import QtRemoteObjects

import sqlite3
from PyQt5.QtSql import *

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QAbstractItemView
import random
from copy import deepcopy

# from PyQt5.QtCore import Qt
# q = QSqlQuery("select * FROM Teacher", con)
# modelForComboBox.setQuery(q)
# modelForComboBox.select()

# tableDel.setCellWidget(y, 0, btnDel)    # для TableWidget
# tableData.setIndexWidget(model.index(y, model.columnCount()-1), btnDel)

# model.index(row, column).data()

# new check-able combo box

# class CheckableComboBox3(QComboBox):
#
#     # constructor
#     def __init__(self, parent=None):
#         super(CheckableComboBox, self).__init__(parent)
#         self.view().pressed.connect(self.handleItemPressed)
#         self.setModel(QStandardItemModel(self))
#
#
#     count = 0
#
#     # action called when item get checked
#     def do_action(self):
#         print(self.count, 'action')
#         # window.label.setText("Checked number : " + str(self.count))
#
#     # when any item get pressed
#     def handleItemPressed(self, index):
#         print('handle', index)
#         # getting the item
#         item = self.model().itemFromIndex(index)
#
#         # checking if item is checked
#         if item.checkState() == Qt.Checked:
#
#             # making it unchecked
#             item.setCheckState(Qt.Unchecked)
#
#         # if not checked
#         else:
#             # making the item checked
#             item.setCheckState(Qt.Checked)
#
#             self.count += 1
#
#             # call the action
#             self.do_action()
#
# class CheckableComboBox2(QtWidgets.QComboBox):
#     # once there is a checkState set, it is rendered
#     # here we assume default Unchecked
#     def addItem(self, text):
#         super(CheckableComboBox, self).addItem(text)
#         print(QtCore.Qt.ItemFlag)
#         item = self.model().item(self.count()-1,0)
#         item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsUserCheckable)
#         item.setCheckState(QtCore.Qt.Unchecked)
#         self.view().pressed.connect(self.action)
#
#     def action(self, index):
#         item = self.model().itemFromIndex(index)
#         if item.checkState() == QtCore.Qt.Checked:
#             item.setCheckState(QtCore.Qt.Unchecked)
#             print('uncheck ', end='')
#         else:
#             print('check ', end='')
#             item.setCheckState(QtCore.Qt.Checked)
#         print('нажато', index)
#
#     def itemChecked(self, index):
#         print('12')
#         item = self.model().item(i,0)
#         return item.checkState() == QtCore.Qt.Checked

from PyQt5.QtWidgets import (QApplication, QWidget, QHBoxLayout,
                             QLineEdit, QListView, QToolButton)
from PyQt5.QtCore import Qt, QAbstractListModel, pyqtSignal, pyqtSlot, QPoint
from PyQt5.QtGui import QIcon, QFont

style_sheet = '''
QListView {
    background-color:white;
    font-size:16px;
    font-family:Comic Sans MS;
}
QListView::item:alternate {
    background:#f7f7f7;
}
QListView::item::hover {
    background: #0ff;
}
'''


class Timer():
    def __init__(self):  # таймер
        self.st = time.time()

    def end(self):
        return float("%.2f" % (time.time() - self.st))


class ListModel(QAbstractListModel):
    def __init__(self):
        super().__init__()
        self._data = []
        self._checklist = []

    def rowCount(self, index):
        return len(self._data)

    def data(self, index, role):
        row = index.row()
        if role == Qt.DisplayRole:
            return self._data[row]
        elif role == Qt.DecorationRole:
            return QIcon('Dark_rc/checkbox_checked.png') if self._checklist[row] else QIcon(
                'Dark_rc/checkbox_unchecked.png')

    def flags(self, index):
        return Qt.ItemIsEnabled  # | Qt.ItemIsSelectable

    def load(self, data_check):
        self.beginResetModel()
        self._data = list(data_check.keys())
        # print(list(data_check.keys()))
        self._checklist = list(data_check.values())
        self.endResetModel()  # ?

    def get(self):
        return ', '.join(x for x, y in zip(self._data, self._checklist) if y)


class View(QListView):
    def __init__(self):
        super().__init__()
        self.setStyleSheet("padding: 0px; margin: 0px;")
        # self.window = QDialog()
        self.setAlternatingRowColors(True)
        self.setStyleSheet(style_sheet)
        self.model = ListModel()
        self.setModel(self.model)


class PopupWidget(QWidget):
    state_changed = pyqtSignal(str)
    change_element = pyqtSignal(str, bool, QWidget)

    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self.setWindowFlags(Qt.Popup)
        self.initView()
        self.view.clicked.connect(self.state_change)  # отслеживается клик по элементу
        self.time_for_open = Timer()

    def state_change(self, index):
        self.view.model._checklist[index.row()] ^= True  # a^b   0^0=0   0^1=1     1^0=1   1^1=0
        self.view.model.layoutChanged.emit()
        self.state_changed.emit(self.view.model.get())
        self.change_element.emit(self.view.model._data[index.row()], self.view.model._checklist[index.row()],
                                 self.widget)

    def initView(self):
        self.view = View()
        self.view.setAlternatingRowColors(True)
        hbox = QHBoxLayout(self)
        hbox.setContentsMargins(1, 1, 1, 1)
        hbox.setSpacing(0)
        hbox.addWidget(self.view)

    def closeEvent(self, event):
        self.time_for_open = Timer()


class MyCheckableComboBox(QWidget):
    def __init__(self):
        super().__init__()
        self.line = QLineEdit()
        self.line.setReadOnly(True)
        self.btn = QToolButton()
        self.setStyleSheet("padding: 0px; margin: 0px;")
        self.btn.setIcon(QIcon('img_rc/array_down.png'))
        self.popup_widget = PopupWidget(self)
        self.changeElement = self.popup_widget.change_element

        hbox = QHBoxLayout(self)
        hbox.setContentsMargins(1, 1, 1, 1)
        hbox.setSpacing(0)
        hbox.addWidget(self.line)
        hbox.addWidget(self.btn)
        self.btn.clicked.connect(self.show_popup)
        self.popup_widget.state_changed.connect(lambda x: self.line.setText(x))

    def addElements(self, elements):  # {data : is_check}
        self.popup_widget.view.model.load(elements)
        self.line.setText(self.popup_widget.view.model.get())

    def show_popup(self):
        if not self.popup_widget.isVisible() and self.popup_widget.time_for_open.end() >= 0.4:  # если сейчас окно закрыто и с момента закрытия прошло 0.4 секунды
            self.popup_widget.setGeometry(100, 200, 100, 100)
            self.popup_widget.resize(self.width(), min(int(27.6 * 6) + 1, 29 * len(
                self.popup_widget.view.model._data)))  # 29 - ширина одного элемента (+-)
            self.popup_widget.move(self.mapToGlobal(QPoint(0, self.height())))
            self.setFont(self.font())
            self.popup_widget.show()


class DataBase():
    def __init__(self):
        self.con = sqlite3.connect('../../Downloads/Timetable.db')
        cur = self.con.cursor()

        cur.execute("""PRAGMA foreign_keys = ON;""")
        cur.execute("""CREATE TABLE IF NOT EXISTS Teacher
        (
          fullName TEXT,
          id INTEGER PRIMARY KEY AUTOINCREMENT
        );
        """)

        cur.execute("""CREATE TABLE IF NOT EXISTS Lesson
        (
          title TEXT,
          hour INT,
          audienceNumber TEXT,
          teacherId INT,
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          FOREIGN KEY (teacherId) REFERENCES Teacher(id) ON DELETE SET NULL ON UPDATE CASCADE
        );
        """)

        cur.execute("""CREATE TABLE IF NOT EXISTS GroupTimetable
        (
          dayOfWeek INT,
          weekNumber INT,
          lessonNumber INT,
          lessonId INT,
          FOREIGN KEY (lessonId) REFERENCES Lesson(id) ON DELETE SET NULL ON UPDATE CASCADE
        );""")

        cur.execute("""CREATE TABLE IF NOT EXISTS Groups
        (
          title TEXT,
          courseNumber INT,
          PRIMARY KEY (title)
        );""")

        cur.execute("""CREATE TABLE IF NOT EXISTS TeacherAttendance
        (
          dayOfWeek INT,
          lessonNumber INT,
          teacherId INT,
          FOREIGN KEY (teacherId) REFERENCES Teacher(id) ON DELETE CASCADE ON UPDATE CASCADE
        );
        """)

        cur.execute("""CREATE TABLE IF NOT EXISTS GroupLesson
        (
          lessonId INT,
          groupTitle TEXT,
          FOREIGN KEY (lessonId) REFERENCES Lesson(id) ON DELETE CASCADE ON UPDATE CASCADE,
          FOREIGN KEY (groupTitle) REFERENCES Groups(title) ON DELETE CASCADE ON UPDATE CASCADE
        );""")
        cur.close()

    def addTeacher(self, teacher):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO Teacher VALUES (?,null)""", (teacher,))
        self.con.commit()

        cur.execute(f"""SELECT id FROM Teacher WHERE fullName = '{teacher}'""")
        for id in cur.fetchall():
            pass
        self.addTeacherTimeWork(id[0])
        cur.close()
        return id[0]

    def addLesson(self, teacher, hours, lesson):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO Lesson VALUES (?,?,?,?,null)""", (lesson, hours, '', teacher))
        self.con.commit()

        cur.execute(f"""SELECT id FROM Lesson WHERE teacherId = '{teacher}'""")
        for id in cur.fetchall():
            pass
        cur.close()
        return id[0]

    def addGroup(self, group, courseNumber):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO Groups VALUES (?,?)""", (group, courseNumber))
        self.con.commit()
        cur.close()

    def getTeacher(self):
        cur = self.con.cursor()
        cur.execute("""SELECT fullName, id FROM Teacher""")
        data = {}
        for rec in cur.fetchall():
            data[rec[1]] = rec[0]
        cur.close()
        return data

    def getTitleTable(self, titleTable, titlePKColumn):
        self.con.row_factory = sqlite3.Row
        cursor = self.con.execute(f'select * from {titleTable}')
        row = cursor.fetchone()
        names = row.keys()
        return names.index(titlePKColumn)

    def getTitleGroup(self):
        cur = self.con.cursor()
        cur.execute("""SELECT title FROM Groups""")
        data = [d[0] for d in cur.fetchall()]
        cur.close()
        return data

    def delRow(self, id, titleTable, titlePKColumn):
        cur = self.con.cursor()
        cur.execute(f"""DELETE FROM {titleTable} WHERE {titlePKColumn} = '{id}'""")
        self.con.commit()
        cur.close()

    def addGroupLesson(self, id_lesson, title_group):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO GroupLesson VALUES (?, ?)""", (id_lesson, title_group))
        self.con.commit()
        cur.close()

    def delGroupLesson(self, id_lesson, title_group):
        cur = self.con.cursor()
        cur.execute(f"""DELETE FROM GroupLesson WHERE lessonId = '{id_lesson}' AND groupTitle = '{title_group}'""")
        self.con.commit()
        cur.close()

    def getGroupLesson(self, idLesson):
        cur = self.con.cursor()
        cur.execute(f"""SELECT groupTitle FROM GroupLesson WHERE lessonId = '{idLesson}' """)
        data = [d[0] for d in cur.fetchall()]
        cur.close()
        return data

    def addTeacherTimeWork(self, id_teacher):
        cur = self.con.cursor()
        for day in range(6):
            for lesson in range(4):
                cur.execute(f"""INSERT INTO TeacherAttendance VALUES (?,?,?)""", (day, lesson, id_teacher))
        self.con.commit()
        cur.close()

    def getTeacherTimeWork(self, id_teacher):
        cur = self.con.cursor()
        timeWork = [[False, False, False, False] for _ in range(6)]
        cur.execute(f"""SELECT dayOfWeek, lessonNumber FROM TeacherAttendance WHERE teacherId = {id_teacher}""")
        for day, lesson in cur.fetchall():
            timeWork[day][lesson] = True
        return timeWork

    def add_work_day(self, day, les, id):
        """удаление и добавление рабочего дня для учителя"""
        cur = self.con.cursor()
        cur.execute("""INSERT INTO TeacherAttendance VALUES (?,?,?)""", (day, les, id))
        self.con.commit()
        cur.close()

    def get_teachers_for_group_in_day(self):
        """для генерации расписания {day:{numLesson:{group:[teachers]}}}"""
        cur = self.con.cursor()
        cur.execute("""SELECT DISTINCT TeacherAttendance.dayOfWeek, TeacherAttendance.lessonNumber, GroupLesson.groupTitle, TeacherAttendance.teacherId
                        FROM GroupLesson, Lesson, TeacherAttendance WHERE TeacherAttendance.teacherId = Lesson.teacherId and GroupLesson.lessonId = Lesson.id""")
        data = {}
        for day, numLesson, group, teacher in cur.fetchall():
            if day not in data.keys():
                data[day] = {}
            if numLesson not in data[day].keys():
                data[day][numLesson] = {}
            if group not in data[day][numLesson].keys():
                data[day][numLesson][group] = []
            data[day][numLesson][group].append(teacher)
        return data

    def get_group_teacher_lesson(self):
        """{group:teacher:[[lessonID, quantityLessonForTwoWeek], ...]}"""
        cur = self.con.cursor()
        cur.execute("""SELECT DISTINCT GroupLesson.groupTitle, Lesson.teacherId, GroupLesson.lessonId, Lesson.hour FROM GroupLesson, Lesson WHERE GroupLesson.lessonId = Lesson.id""")
        query = cur.fetchall()
        data = {}
        group_hours = self.get_hours_group()
        for group, teacher, lesson, hour in query:
            if group not in data.keys():
                data[group] = {}
                group_hours[group] //= 36   # получаем кол-во учебных недель
            if teacher not in data[group].keys():
                data[group][teacher] = []
            data[group][teacher].append([lesson, hour//group_hours[group]])
        return data, group_hours

    def get_hours_group(self):
        """общее кол-во часов для группы за семестр"""
        cur = self.con.cursor()
        cur.execute("""SELECT GroupLesson.groupTitle, SUM(Lesson.hour) FROM GroupLesson, Lesson WHERE Lesson.id = GroupLesson.lessonId GROUP BY GroupLesson.groupTitle""")
        group_hours = {}
        for group, hours in cur.fetchall():
            group_hours[group] = hours
        cur.close()
        return group_hours

    def getAllTeacherTimeWork(self):
        """получение всех возможных дней работы учителей"""
        cur = self.con.cursor()
        cur.execute("""SELECT id FROM Teacher""")
        teacherID_TimeWork = {}
        for id in cur.fetchall():
            id = id[0]
            teacherID_TimeWork[id] = [self.getTeacherTimeWork(id), self.getTeacherTimeWork(id)]

        cur.close()
        return teacherID_TimeWork

    def get_group_lesson(self):
        cur = self.con.cursor()
        cur.execute("""SELECT Lesson.hour, lessonId, groupTitle FROM GroupLesson, Lesson WHERE GroupLesson.lessonId = Lesson.id""")
        number_free_lessons = 12
        group_lesson = {}
        group_hours = self.get_hours_group()
        # print(group_hours)
        for hour, lesson, group in cur.fetchall():
            if group not in group_lesson.keys():
                group_lesson[group] = [None]*number_free_lessons
            group_lesson[group] += [lesson]*(hour//(group_hours[group]//36))
            # print([lesson]*(hour//(group_hours[group]//36)))
            # print(lesson, hour//(group_hours[group]//36), hour, group, group_hours[group])
        return group_lesson

    def insert_time_table(self, timeTable):
        cur = self.con.cursor()
        cur.execute("""DELETE FROM GroupTimetable""")
        # weekNumber dayOfWeek lessonNumber lessonId
        for group_lesson in timeTable:
            for week in range(2):
                for day in range(6):
                    for num_lesson in range(4):
                        num = week*6*4 + day*4 + num_lesson
                        lesson = group_lesson[num]
                        if lesson != None:
                            cur.execute("""INSERT INTO GroupTimetable VALUES (?,?,?,?)""", (week, day, num_lesson, lesson))
        cur.close()
        self.con.commit()

    def getTimeTable(self):
        cur = self.con.cursor()
        cur.execute("""SELECT groupTitle, weekNumber, dayOfWeek, lessonNumber, lessonId FROM GroupTimetable, GroupLesson WHERE GroupLesson.lessonId = GroupTimetable.lessonId""")
        TimeTable = {}
        for group, week, day, num_lesson, lesson_id in cur.fetchall():
            if group not in TimeTable.keys():
                TimeTable[group] = {}
            if week not in TimeTable[group].keys():
                TimeTable[group][week] = {}
            if day not in TimeTable[group][week].keys():
                TimeTable[group][week][day] = {}
            TimeTable[group][week][day][num_lesson] = lesson_id
        return TimeTable


class table(QtWidgets.QWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.layout = QtWidgets.QGridLayout()

        self.setLayout(self.layout)

        self.tableData = QtWidgets.QTableView()
        self.layout.addWidget(self.tableData)

        con = QSqlDatabase.addDatabase("QSQLITE")
        con.setDatabaseName('Timetable.db')
        con.open()

        model = QSqlTableModel()
        model.setTable('Teacher')
        model.select()
        self.tableData.setModel(model)

        self.tableData.setTextElideMode(Qt.ElideNone)
        self.tableData.horizontalHeader().setStretchLastSection(True)
        self.tableData.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tableData.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tableData.setSortingEnabled(True)
        # model.dataChanged.connect(self.changeTable)
        # self.tableData.horizontalHeader().setStretchLastSection(True)
        # self.tableData.resizeColumnToContents(0)


class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.DB = DataBase()
        self.ui = PyQt5.uic.loadUi("main.ui")
        setStyle(self.ui)
        self.ui.show()
        # self.setFont(QFont("Comics Sans MS", 10))
        # self.clickBtnAdd = False
        # self.xclx = ParsingXLSX(self.DB)
        teacher = tableWithEditing(['Учителя', 'id', '', ''], "Teacher", 'id', self.DB, extra_button=True)
        self.ui.horizontalLayout.addWidget(teacher)
        group = tableWithEditing(['Группы', 'Курс', ''], "Groups", 'title', self.DB)
        self.ui.horizontalLayout_2.addWidget(group)
        lesson = tableWithEditing(['Занятие', 'Часы', 'Кабинет', 'Учитель', 'id', 'Группа', ''], "Lesson", 'id',
                                  self.DB, self.ui.tabWidget)
        self.ui.horizontalLayout_3.addWidget(lesson)
        timeTable = TimeTable(self.DB, self.ui)
        # layout = QtWidgets.QGridLayout()
        # TimeTableTab = TimeTable(self.DB)
        # self.ui.tabWidget.addTab(QWidget(), "Расписание")
        # self.ui.tabWidget.
        # QtWidgets.QTabWidget.addTab()
        # GenerateTimeTable(self.DB)
        # GeneticGenerationTimeTable(self.DB)

class TimeTable(QWidget):
    def __init__(self, DB, ui):
        super().__init__()
        self.DB = DB
        self.table = ui.TimeTable

    def load_TimeTable(self):
        TimeTable = self.DB.getTimeTable()
        # numrows = len(data)  # 6 rows in your example
        # numcols = len(titels)  # 3 columns in your example
        # self.ui.tableWidget.setColumnCount(numcols)
        # self.ui.tableWidget.setRowCount(numrows)
        # self.ui.tableWidget.setHorizontalHeaderLabels(titels)
        #
        # for row in range(numrows):
        #     for column in range(numcols):
        #         self.ui.tableWidget.setItem(row, column, QTableWidgetItem((data[row][column])))
        # self.ui.tableWidget.horizontalHeader().setStretchLastSection(True)
        # self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)



class tableWithEditing(QtWidgets.QWidget):
    def __init__(self, headTable, titleTable, titlePKColumn, DB, tabWidget=False, extra_button=False, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.layout = QtWidgets.QGridLayout()
        self.titlePKColumn = titlePKColumn  # для удаления столбца
        self.setLayout(self.layout)
        self.titleTable = titleTable
        self.tableData = QtWidgets.QTableView()
        self.tableData.setSortingEnabled(True)
        self.layout.addWidget(self.tableData)

        con = QSqlDatabase.addDatabase("QSQLITE")
        con.setDatabaseName('Timetable.db')
        con.open()

        model = QSqlTableModel()
        model.setTable(self.titleTable)
        model.select()
        self.model = model
        self.tableData.setModel(model)
        self.model = self.setTableData()

        self.headTable = headTable
        self.DB = DB
        self.clickBtnAdd = False
        self.extra_button = extra_button

        self.setBtnAdd()
        self.setWidgets(title=True)
        self.model.dataChanged.connect(self.changeTable)
        # self.changeTable()
        if tabWidget:
            tabWidget.currentChanged.connect(self.changeTab)
            self.tabWidget = tabWidget
        # self.tableData.setTextElideMode(Qt.ElideNone)
        self.tableData.horizontalHeader().setStretchLastSection(True)
        self.tableData.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tableData.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.tableData.horizontalHeader().setSectionResizeMode(self.model.columnCount()-2, 1)
        # self.tableData.horizontalHeader().setSectionResizeMode(self.model.columnCount()-1, 1)

    def setWidgets(self, title=False):
        """
        после .select() нужно вставить виджеты и скрыть колонку id
        :param title: при первом заполнении таблицы так же нужно вставить заголовки
        :return:
        """
        self.model.insertColumn(self.model.columnCount())  # для кнопок удаления
        if self.titleTable == "Teacher" or self.titleTable == "Lesson":
            self.tableData.setColumnHidden(self.model.columnCount() - 2,
                                           True)  # спрятать id (предпоследняя колонна) для Teacher or Lesson
        if self.titleTable == "Groups":
            self.setSpinBox()
        elif self.titleTable == "Lesson":
            self.setSpinBox(1, 99999)  # добавляем spinbox
            self.setComboBox(3, self.DB.getTeacher())
            self.setCheckableComboBox(self.DB.getTitleGroup())

        if self.extra_button:
            self.setExtraButton()  # используем столбец для кнопки удаления
            self.model.insertColumn(self.model.columnCount())  # и добавляем новый столбец для кнопки удаления

        self.setBtnDel()
        if title:
            for i, head in enumerate(self.headTable):
                self.model.setHeaderData(i, Qt.Horizontal, head)
        self.changeTable()
        self.setStyleSheet("""
        QPushButton { text-align: center;}
        QLabel {background-color: rgb(255,128,0);}
        QWidget {font-size:16px; font-family:Comic Sans MS;}
        """)

    def setExtraButton(self):
        row = self.model.rowCount()  # кол-во строк
        self.extraBtn_row = {}
        for y in range(row):
            extra_btn = QPushButton()
            size = 37
            extra_btn.setMinimumSize(size, size)
            extra_btn.setMaximumSize(size, size)  # x, y
            icon = QIcon("img/teach_negative.png")
            extra_btn.setIconSize(QtCore.QSize(size - 7, size - 7))
            extra_btn.setIcon(icon)
            cell_widget = QWidget()
            lay_out = QHBoxLayout(cell_widget)
            lay_out.addWidget(extra_btn)
            lay_out.setAlignment(QtCore.Qt.AlignCenter)
            lay_out.setContentsMargins(0, 0, 0, 0)  # Рисуем границы
            cell_widget.setLayout(lay_out)
            self.extraBtn_row[extra_btn] = y
            self.tableData.setIndexWidget(self.model.index(y, self.model.columnCount() - 1), cell_widget)
            extra_btn.clicked.connect(lambda: self.openWorkingTime())

    def openWorkingTime(self):
        btn = self.sender()
        row = self.extraBtn_row[btn]
        # print(row, id_Teacher := self.model.index(row, 0).data(), id_Teacher := self.model.index(row, 1).data())
        # id_Teacher = self.model.index(0, row).data()
        # id_Teacher = self.model.index(1, row).data()
        id_Teacher = self.model.index(row, 1).data()
        win = WorkingTime(name=self.model.index(row, 0).data(), id=id_Teacher, DB=self.DB)
        # setStyle(win)
        win.show()
        win.exec()

    def setFromCheckableComboBox(self, el, is_check, comboBox):
        row = self.сheckableСomboBox_row[comboBox]
        id_lesson = self.model.index(row, 4).data()
        if is_check:
            self.DB.addGroupLesson(id_lesson, el)
        else:
            self.DB.delGroupLesson(id_lesson, el)

    def setCheckableComboBox(self, data):
        self.model.insertColumn(self.model.columnCount())
        column = self.model.columnCount() - 2
        self.сheckableСomboBox_row = {}
        for row in range(self.model.rowCount()):
            data_check = {d: False for d in data}
            d_cheked = self.DB.getGroupLesson(self.model.index(row, 4).data())
            for d in d_cheked:
                data_check[d] = True
            сheckableСomboBox = MyCheckableComboBox()
            сheckableСomboBox.changeElement.connect(
                lambda el, is_check, comboBox: self.setFromCheckableComboBox(el, is_check, comboBox))
            сheckableСomboBox.addElements(data_check)
            # сheckableСomboBox.addItems(data)
            # сheckableСomboBox.lineEdit().setText('Выберите что-то')
            # nowId = self.model.index(row, column).data()  # берем из таблицы внешней ключ
            # if nowId == '' and data:
            # self.updateRow(row, column, int(сheckableСomboBox.currentText().split('|')[1]))
            # comboBox.setStyleSheet("""background-color:rgb(255,128,138);""")
            # elif nowId != '':
            #     сheckableСomboBox.setCurrentText(data[nowId]+' | '+str(nowId))
            self.tableData.setIndexWidget(self.model.index(row, column), сheckableСomboBox)
            self.сheckableСomboBox_row[сheckableСomboBox] = row
            # сheckableСomboBox.currentIndexChanged.connect(self.setFromComboBox)
            # сheckableСomboBox.popup_widget.view.model._checklist[0] = True
        # print(self.сheckableСomboBox_row)
        # print(self.checkableComboBox_row)     # в чем отличие от прошлой строчки??????

    def addRowTable(self):
        """добавление строки в таблицу при нажатии на кнопку"""
        # print('123')
        self.clickBtnAdd = True
        self.model.submitAll()  # ячейка при добавлении новой может еще редактироваться пользователем, поэтому надо сначала принять введенное (при редактировании нельзя добавить)
        record = self.model.record()
        if self.titleTable == "Teacher":
            record.remove(record.indexOf("id"))
            record.setValue("fullName", "")
        elif self.titleTable == "Groups":
            record.setValue("title", "")
            record.setValue("courseNumber", 1)
        elif self.titleTable == "Lesson":
            record.remove(record.indexOf("id"))
            record.setValue("title", "")
            record.setValue("hour", 1)
            record.setValue("audienceNumber", 1)
            record.setValue("teacherId", "")
        self.model.insertRecord(-1, record)

        self.model.select()
        self.setWidgets()
        if self.titleTable == "Teacher":
            id_teacher = self.model.index(self.model.rowCount() - 1, 1).data()
            self.DB.addTeacherTimeWork(id_teacher)

        self.tableData.scrollToBottom()  # пролистываем вниз
        self.clickBtnAdd = False

    def setBtnAdd(self):
        """добавляем кнопку добавления новой строки"""
        btnAdd = QtWidgets.QPushButton()
        self.layout.addWidget(btnAdd)
        btnAdd.clicked.connect(self.addRowTable)
        btnAdd.setStyleSheet("""background-color: rgb(255,255,255); text-align: center;""")
        size = 30
        btnAdd.setMinimumSize(size, size)
        btnAdd.setIconSize(QtCore.QSize(size - 7, size - 7))

        icon = QIcon("img/add.png")
        btnAdd.setIcon(icon)

    def delRowTable(self):
        """удаление строки из таблицы и бд при нажатии на кнопку"""
        btnDel = self.sender()
        row = self.tableData.indexAt(btnDel.pos()).row()
        column = self.DB.getTitleTable(self.titleTable, self.titlePKColumn)  # по названию возвращаем номер столбца
        dataId = self.model.index(row, column).data()
        self.DB.delRow(dataId, self.titleTable, self.titlePKColumn)  # из-за каскадного удаления
        self.model.submitAll()
        self.model.select()
        self.setWidgets()

    def setBtnDel(self):
        """установка кнопок удаления в таблицу"""
        row = self.model.rowCount()  # кол-во строк
        for y in range(row):
            btnDel = QPushButton()
            size = 37
            btnDel.setMinimumSize(size, size)
            btnDel.setMaximumSize(size, size)  # x, y
            icon = QIcon("img/del.png")
            btnDel.setIcon(icon)
            btnDel.setIconSize(QtCore.QSize(size - 17, size - 17))
            # btnDel.setStyleSheet("""padding: 0px 230px 0px;""")
            self.tableData.setIndexWidget(self.model.index(y, self.model.columnCount() - 1), btnDel)
            btnDel.clicked.connect(lambda: self.delRowTable())

    def setSpinBox(self, column=1, max=5):
        @pyqtSlot()
        def setFromSpinBox(tableData):  # вызывается при установке значения из spinBox в БД
            self.model.submitAll()
            spinBox = self.sender()
            rowSpinBox = tableData.indexAt(spinBox.pos()).row()
            columnSpinBox = tableData.indexAt(spinBox.pos()).column()
            numForSet = spinBox.value()  # вставляем число в бд
            record = self.model.record()
            for col in range(self.model.columnCount() - 1):
                record.setValue(col, self.model.index(rowSpinBox, col).data())
                if col == columnSpinBox:
                    record.setValue(1, numForSet)
            self.model.updateRowInTable(rowSpinBox, record)

        for row in range(self.model.rowCount()):  # кол-во строк
            spinBox = QSpinBox()
            spinBox.wheelEvent = lambda event: None
            spinBox.setMaximum(max)
            spinBox.setMinimum(1)
            spinBox.setValue(self.model.index(row, column).data())  # вставляем в spinbox данные из ячейки
            self.tableData.setIndexWidget(self.model.index(row, column), spinBox)
            spinBox.editingFinished.connect(lambda: setFromSpinBox(self.tableData))

    def setFromComboBox(self):
        '''при изменении значении в combobox вставляем это значение в БД'''
        # print('комбобокс изменен')
        comboBox = self.sender()
        comboBox.setStyleSheet("""""")
        rowComboBox = self.tableData.indexAt(comboBox.pos()).row()
        columnComboBox = self.tableData.indexAt(comboBox.pos()).column()
        id = int(comboBox.currentText().split('|')[1])
        self.updateRow(rowComboBox, columnComboBox, id)

    def setComboBox(self, column, data):
        """
        вставка QCombobox в таблицу
        :param column: колонка куда вставляется QComboBox
        :param data: данные для вставки {id:data}
        """
        for row in range(self.model.rowCount()):
            comboBox = QComboBox()
            comboBox.addItems([v + ' | ' + str(k) for k, v in data.items()])
            nowId = self.model.index(row, column).data()  # берем из таблицы внешней ключ
            if nowId == '' and data:
                self.updateRow(row, column, int(comboBox.currentText().split('|')[1]))
                comboBox.setStyleSheet("""background-color:rgb(255,128,138);""")
            elif nowId != '':
                comboBox.setCurrentText(data[nowId] + ' | ' + str(nowId))
            self.tableData.setIndexWidget(self.model.index(row, column), comboBox)
            comboBox.currentIndexChanged.connect(self.setFromComboBox)

            comboBox.setEditable(True)  # для поиска в comboBox
            comboBox.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
            # change completion mode of the default completer from InlineCompletion to PopupCompletion
            comboBox.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)

    def updateRow(self, row, column, id):
        """
        обнавляем строку в таблице (при добавлении значения в бд из виджета)
        :param row:
        :param id:
        :return:
        """
        record = self.model.record()
        for col in range(self.model.columnCount() - 1):
            print(self.model.index(row, col).data())
            if col == column:
                record.setValue(col, id)
            else:
                record.setValue(col, self.model.index(row, col).data())
        self.model.updateRowInTable(row, record)

    def changeTable(self):
        """вызывается при изменении ячейки"""
        # print('ячейка изменена')
        # self.tableData.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        # self.tableData.horizontalHeader().setSectionResizeMode(self.model.columnCount()-2, 0)
        # # self.tableData.horizontalHeader().setStretchLastSection(True)
        # self.tableData.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.tableData.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        if not self.clickBtnAdd:  # я не знаю почему, но если вызывать submitAll при добавлении строки, то программа вылетает, но submitAll нужно вызывать, чтобы сразу же сохранить изменения в ячейки таблицы, а не после нажатия на enter
            self.model.submitAll()

    def changeTab(self):
        """
        обновляется таблица для обновления данных в comboBox
        вызывается при переходе на вкладку 'Занятия'
        :return:
        """
        print('вкладка изменена на', self.tabWidget.currentIndex())
        # if self.titleTable == "Teacher" and self.ui.tabWidget.currentIndex() == 0:
        # elif self.titleTable == "Groups" and self.ui.tabWidget.currentIndex() == :
        if self.titleTable == "Lesson" and self.tabWidget.currentIndex() == 2:
            self.model.select()
            self.setWidgets()

    def setTableData(self):
        """установка данных в таблицу из БД """
        con = QSqlDatabase.addDatabase("QSQLITE")
        con.setDatabaseName('Timetable.db')
        con.open()

        model = QSqlTableModel()
        model.setTable(self.titleTable)
        model.select()

        self.tableData.setModel(model)
        # self.tableData.verticalHeader().hide()   # спрятать цифры сбоку

        # self.setWidgets(self.model)
        # self.model.insertColumn(self.model.columnCount())    # для удаления
        # if self.titleTable == "Teacher" or self.titleTable == "Lesson":
        #     self.tableData.setColumnHidden(self.model.columnCount()-2, True)  # спрятать id (предпоследняя колонна) для Teacher or Lesson
        # # if self.titleTable == "Lesson":
        # #     self.model.insertColumn(self.model.columnCount()-1)    # для групп

        # for i, head in enumerate(self.headTable):
        #     self.model.setHeaderData(i, Qt.Horizontal, head)
        return model


class WorkingTime(QDialog):
    def __init__(self, name, id, DB):
        super().__init__()
        self.DB = DB
        self.attendance = self.DB.getTeacherTimeWork(id)
        self.id_teacher = id
        self.setWindowTitle('Расписание учителя')
        self.setWindowIcon(QIcon("img/teacher.png"))

        self.resize(650, 380)
        lay = QVBoxLayout()
        self.setLayout(lay)
        text = QLabel()
        text.setText('Расписание учителя <<' + name + '>>')
        text.setAlignment(QtCore.Qt.AlignCenter)
        lay.addWidget(text)
        self.table = QTableWidget(8, 6)
        # self.table.setColumnCount(6)
        # self.table.setRowCount(8)
        lay.addWidget(self.table)
        self.table.cellClicked.connect(lambda row, column: self.click_work_time(row, column))
        self.table.setHorizontalHeaderLabels(['Понедельник', "Вторник", "Среда", "Четверг", "Пятница", "Суббота"])
        self.table.setVerticalHeaderLabels(['1', '1', '2', '2', '3', '3', '4', '4'])
        self.table.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.red = [250, 10, 10]
        self.dark_red = [199, 0, 0]
        self.green = [0, 255, 0]
        self.dark_green = [0, 179, 0]
        self.fill_table_teacher_attendance()

        self.table.setSpan(0, 0, 2, 1)
        self.table.setSpan(1, 0, 4, 4)

        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        for row in range(6):
            self.table.horizontalHeader().setSectionResizeMode(row, 1)
        for col in range(8):
            self.table.verticalHeader().setSectionResizeMode(col, 1)

        self.table.setMouseTracking(True)
        self.current_hover = [0, 0]
        self.table.cellEntered.connect(self.cellHover)
        self.table.setSelectionMode(False)

    def cellHover(self, row, column):   # при наведении
        old_item = self.table.item(self.current_hover[0], self.current_hover[1])
        if self.current_hover != [row, column]:
            day, lesson = old_item.column(), old_item.row() // 2
            if self.attendance[day][lesson]:  # работает --> зеленый
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.green))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.green))
            else:  # не работает --> красный
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.red))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.red))

            day, lesson = column, row // 2
            if self.attendance[day][lesson]:  # работает --> зеленый
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.dark_green))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.dark_green))
            else:  # не работает --> красный
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.dark_red))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.dark_red))
        self.current_hover = [row, column]

    def fill_table_teacher_attendance(self):    # закрасить при открытии окна
        attendance = self.attendance
        for day in range(6):
            for lesson in range(4):
                self.table.setItem(lesson * 2, day, QTableWidgetItem())
                self.table.setItem(lesson * 2 + 1, day, QTableWidgetItem())
                if attendance[day][lesson]:
                    self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.green))
                    self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.green))
                else:
                    self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.red))
                    self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.red))

    def click_work_time(self, row, column):
        attendance = self.attendance
        day, lesson = column, row // 2
        attendance[day][lesson] ^= True

        if attendance[day][lesson]:  # работает --> зеленый
            self.DB.add_work_day(day, lesson, self.id_teacher)    # добавляем в бд
            self.setStyleSheet("""QTreeView,QListView,QTableView{
                border: 3px solid #B8B8B8;
            }
            """)
            self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.green))
            self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.green))
        else:  # не работает --> красный
            # self.DB.change_work_day(f"""DELETE FROM TeacherAttendance WHERE dayOfWeek = {day} and lessonNumber = {lesson} and teacherId = {self.id_teacher}""")   # добавляем в бд
            self.setStyleSheet("""QTreeView,QListView,QTableView{
                            border: 3px solid #B8B8B8;
                            }""")
            self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.red))
            self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.red))


import openpyxl
class ParsingXLSX():
    def __init__(self, DB):
        super().__init__()
        self.DB = DB
        file_name = 'timeTable.xlsx'
        self.file = openpyxl.load_workbook(file_name)
        titles_lists = self.file.sheetnames
        # print(titles_lists)
        # self.teacher_id = {}
        self.teacher_group_lesson = {}
        self.get_teacher_from_xlsx()
        # print(*self.teacher_group_lesson.items(), sep='\n')
        self.save_in_DB()

    def get_teacher_from_xlsx(self, start_row=9, col_lesson = 2, col_techaer=3, col_hours_1=9, col_hours_2=10):
        for group in self.file.sheetnames:
            now_row = start_row
            work_list = self.file[group]
            while work_list.cell(row = now_row, column = col_techaer).value != None:
                hours = 0
                hours_1 = work_list.cell(row = now_row, column = col_hours_1).value
                hours_2 = work_list.cell(row = now_row, column = col_hours_2).value
                hours += int(hours_1) if hours_1 else 0
                hours += int(hours_2) if hours_2 else 0
                if hours != 0:
                    lesson = work_list.cell(row = now_row, column = col_lesson).value
                    teacher = work_list.cell(row = now_row, column = col_techaer).value
                    try:
                        self.teacher_group_lesson[teacher][group].append([hours, lesson])
                    except KeyError:    # если у учителя еще нет такой группы
                        try:
                            self.teacher_group_lesson[teacher][group] = [[hours, lesson]]
                        except KeyError:    # если еще не было такого учителя
                            self.teacher_group_lesson[teacher] = {}
                            self.teacher_group_lesson[teacher][group] = [[hours, lesson]]
                now_row += 1

    def save_in_DB(self):
        teacher_id = {}
        is_group = {}
        for teacher in self.teacher_group_lesson.keys():
            id = self.DB.addTeacher(teacher)
            # print(id, teacher)
            # teacher_id[teacher] = id
            for group in self.teacher_group_lesson[teacher]:
                try:
                    if is_group[group]:
                        pass
                except:
                    is_group[group] = True
                    for course in group:
                        if course.isdigit():
                            break
                    self.DB.addGroup(group, int(course))
                for hours, lesson in self.teacher_group_lesson[teacher][group]:
                    id_lesson = self.DB.addLesson(id, hours, lesson)
                    self.DB.addGroupLesson(id_lesson, group)


class GenerateTimeTable():
    def __init__(self, DB):
        """берет значения из бд: все группы с кол-вом уч. недель + для каждой пары возможные учителя для данной группы + для каждого учителя возможные предметы для данной группы + кол-во занятий данного предмета для данной группы"""
        self.DB = DB

        self.teacherID_TimeWork = self.DB.getAllTeacherTimeWork()   # {teacherID:{[day:{numLesson:True|False}}, day:{numLesson:True|False}}]} - 2 недели
        self.teachers_for_group_in_day = self.DB.get_teachers_for_group_in_day()     # {day:{numLesson:{group:[teachers]}}}
        print(*self.teachers_for_group_in_day.items(), sep='\n')
        self.count = 0
        self.group_teacher_lesson, self.all_hours_group = self.DB.get_group_teacher_lesson() # {group:{TeacherID:[[idLesson, кол-во занятий], ...]}} ; {группа:кол-во уч. недель}
        self.all_group = list(self.all_hours_group.keys())  # названия групп
        print(self.group_teacher_lesson)
        self.changesTimeTable = []  # стек изменений расписания
        self.TimeTable = {}
        self.getEmpatyTimeTable()
        self.fill_group_week_quantityLesson()
        # self.teacherID_TimeWork[1][0][0][0] = 123
        print(self.teacherID_TimeWork)
        # print(self.TimeTable)
        self.placement_of_lesson = [[0,2], [1,3], [0,1], [0,3],  [2,3], [1,2]]  # [start,end]
        t = Timer()
        if self.combinationsTimeTable(0, 0, -1, 0, 0, 1):
            print(self.TimeTable)
        else:
            print('не удалось найти нужное расписание')
        print(self.count)
        t.end()

    def fill_group_week_quantityLesson(self, quantityLesson = 18):
        """изначально у каждой группы на каждую неделю выделено по 18 пар"""
        self.group_week_quantityLesson = {}
        for group in self.all_group:
            self.group_week_quantityLesson[group] = [quantityLesson, quantityLesson]

    def getEmpatyTimeTable(self):
        for group in self.all_hours_group.keys():
            self.TimeTable[group] = []
            for week in range(2):
                self.TimeTable[group].append([])
                for day in range(6):
                    self.TimeTable[group][week].append([False]*4)

    def combinationsTimeTable(self, week, day, numGroup, numStartLesson, numEndLesson, nowNumLesson):
        self.count += 1
        if nowNumLesson == numEndLesson+1:
            print(week, day, numGroup)
            if day == 5 and numGroup + 1 == len(self.all_group):    # следующая неделя
                week += 1
                day = 0
                numGroup = 0
            elif numGroup + 1 == len(self.all_group):   # следующий день
                day += 1
                numGroup = 0
            else:   # следующая группа
                numGroup += 1

            if week == 2:
                print(self.TimeTable)
                print(self.count)
                time.sleep(1000)
                return True
            f = False
            for startLesson, endLesson in self.placement_of_lesson:
                self.group_week_quantityLesson[self.all_group[numGroup]][week] -= (endLesson-startLesson+1)
                remaining_lesson = self.group_week_quantityLesson[self.all_group[numGroup]][week]
                remaining_day = 5 - day
                if (remaining_day == 0 and remaining_lesson == 0) or (remaining_day != 0 and remaining_lesson // remaining_day >= 2 and remaining_lesson // remaining_day <= 4):
                    f = True
                    if self.combinationsTimeTable(week, day, numGroup, startLesson, endLesson, startLesson):
                        return True
                self.group_week_quantityLesson[self.all_group[numGroup]][week] += (endLesson-startLesson+1)

            if not f:
                print('ошибка: нет подходящих расстановок пар')

        else:
            for teacher in self.teachers_for_group_in_day[day][nowNumLesson][self.all_group[numGroup]]:
                if self.teacherID_TimeWork[teacher][week][day][nowNumLesson]:
                    self.teacherID_TimeWork[teacher][week][day][nowNumLesson] = False
                    for i, lessson in enumerate(self.group_teacher_lesson[self.all_group[numGroup]][teacher]):  # перебираем доступные занятия для данного учителя
                        if self.group_teacher_lesson[self.all_group[numGroup]][teacher][i][1] != 0:
                            self.TimeTable[self.all_group[numGroup]][week][day][nowNumLesson] = lessson[0]
                            self.group_teacher_lesson[self.all_group[numGroup]][teacher][i][1] -= 1
                            if self.combinationsTimeTable(week, day, numGroup, numStartLesson, numEndLesson, nowNumLesson+1):
                                return True
                            self.group_teacher_lesson[self.all_group[numGroup]][teacher][i][1] += 1
                            self.TimeTable[self.all_group[numGroup]][week][day][nowNumLesson] = False
                    self.teacherID_TimeWork[teacher][week][day][nowNumLesson] = True
                    # if week == 1 and day == 5 and numGroup == len(self.all_hours_group.keys())-1 and numLesson == 3:
        #     return


class GeneticGenerationTimeTable():
    def __init__(self, DB):
        """запросы к бд + 1 поколение"""
        self.DB = DB
        self.POPULATION_SIZE = 500
        self.P_CROSSOVER = 0.9
        self.P_MUTAYION = 0.1
        self.MAX_TIME = 5*60

        t = Timer()
        self.generate_first_generation()
        print(t.end())

    def generate_first_generation(self):
        group_lesson = self.DB.get_group_lesson()
        # print(group_lesson)
        self.Generation_TimeTables = []

        for i in range(self.POPULATION_SIZE):
            self.Generation_TimeTables.append(deepcopy(group_lesson))
            for group in self.Generation_TimeTables[-1]:
                random.shuffle(self.Generation_TimeTables[i][group])
        # print(group, *(self.Generation_TimeTables[i-1].items()), sep='\n')
        # print()
        # print(group, *(self.Generation_TimeTables[i].items()), sep='\n')
        # print(*self.Generation_TimeTables, sep='\n')



def setStyle(ui):
    """Стиль не для всей программы"""
    ui.setStyleSheet("""
    QPushButton { text-align: center;}
        QLabel {background-color: rgb(255,128,0);}
        QWidget {font-size:16px; font-family:Comic Sans MS;
    }

    QTableView::item:selected, QListView::item:selected, QTreeView::item:selected {
        color: #F0F0F0;
        background: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0 #5d5e61, stop:1 #adadad);
    }

    QTableView::item:hover, QListView::item:hover, QTreeView::item:hover {
        color: #F0F0F0;
        background: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0 #B8B8B8, stop:1 #D6D6D6);
    }

    QTreeView,QListView,QTableView{
        border: 1px solid #B8B8B8;
        selection-background-color: #454648;
        selection-color: #F0F0F0;
    }
    """)


if __name__ == '__main__':
    # num = {1:['qwe'], 2:['asd']}
    # a = num.copy()
    # b = num.copy()
    # b[1][0]= '123'
    # print(num, a,b)
    # from copy import deepcopy
    # l  = [[1,2], [3,4]]
    # b = deepcopy(l)
    # b[0][0] = -1
    # print(l, b)

    sys.setrecursionlimit(4000)
    app = QtWidgets.QApplication([])
    f = open('style_gray.qss')
    style = f.read()
    app.setStyleSheet(style)
    # app.setStyleSheet(PyQt5_stylesheets.load_stylesheet_pyqt5(style="style_gray"))
    f.close()

    win = MainWindow()
    sys.exit(app.exec_())