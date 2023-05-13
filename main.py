import sys
import time
import typing
from turtle import color

import PyQt5
from PyQt5 import uic
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTimer
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtGui import QBrush
from PyQt5.QtGui import QFont, QPalette, QColor
from PyQt5.QtGui import QIcon
from PyQt5.QtGui import QStandardItem
from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtWidgets import *
from itertools import product
from PyQt5 import Qt
from PyQt5 import QtRemoteObjects

import sqlite3
from PyQt5.QtSql import *

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QAbstractItemView
import random
from copy import deepcopy
from genAlgorithm import GA
from PyQt5.QtWidgets import (QApplication, QWidget, QHBoxLayout,
                             QLineEdit, QListView, QToolButton)
from PyQt5.QtCore import Qt, QAbstractListModel, pyqtSignal, pyqtSlot, QPoint
from PyQt5.QtGui import QIcon, QFont
from xlsxwriter.workbook import Workbook                         # pip install XlsxWriter
from openpyxl import load_workbook
import openpyxl

style_sheet = '''
QListView {
    background-color:white;
    font-size:16px;
    font-family:Comic Sans MS;
}
QListView::item:alternate {
    background:#f7f7f7;
}
QListView::item::hover {
    background: #0ff;
}
'''


class Timer():
    def __init__(self):  # таймер
        self.st = time.time()

    def end(self):
        return float("%.2f" % (time.time() - self.st))


class ListModel(QAbstractListModel):
    def __init__(self):
        super().__init__()
        self._data = []
        self._checklist = []

    def rowCount(self, index):
        return len(self._data)

    def data(self, index, role):
        row = index.row()
        if role == Qt.DisplayRole:
            return self._data[row]
        elif role == Qt.DecorationRole:
            return QIcon('Dark_rc/checkbox_checked.png') if self._checklist[row] else QIcon(
                'Dark_rc/checkbox_unchecked.png')

    def flags(self, index):
        return Qt.ItemIsEnabled  # | Qt.ItemIsSelectable

    def load(self, data_check):
        self.beginResetModel()
        self._data = list(data_check.keys())
        # print(list(data_check.keys()))
        self._checklist = list(data_check.values())
        self.endResetModel()  # ?

    def get(self):
        return ', '.join(x for x, y in zip(self._data, self._checklist) if y)


class View(QListView):
    def __init__(self):
        super().__init__()
        self.setStyleSheet("padding: 0px; margin: 0px;")
        # self.window = QDialog()
        self.setAlternatingRowColors(True)
        self.setStyleSheet(style_sheet)
        self.model = ListModel()
        self.setModel(self.model)


class PopupWidget(QWidget):
    state_changed = pyqtSignal(str)
    change_element = pyqtSignal(str, bool, QWidget)

    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self.setWindowFlags(Qt.Popup)
        self.initView()
        self.view.clicked.connect(self.state_change)  # отслеживается клик по элементу
        self.time_for_open = Timer()

    def state_change(self, index):
        self.view.model._checklist[index.row()] ^= True  # a^b   0^0=0   0^1=1     1^0=1   1^1=0
        self.view.model.layoutChanged.emit()
        self.state_changed.emit(self.view.model.get())
        self.change_element.emit(self.view.model._data[index.row()], self.view.model._checklist[index.row()],
                                 self.widget)

    def initView(self):
        self.view = View()
        self.view.setAlternatingRowColors(True)
        hbox = QHBoxLayout(self)
        hbox.setContentsMargins(1, 1, 1, 1)
        hbox.setSpacing(0)
        hbox.addWidget(self.view)

    def closeEvent(self, event):
        self.time_for_open = Timer()


class MyCheckableComboBox(QWidget):
    def __init__(self):
        super().__init__()
        self.line = QLineEdit()
        self.line.setReadOnly(True)
        self.btn = QToolButton()
        self.setStyleSheet("padding: 0px; margin: 0px;")
        self.btn.setIcon(QIcon('img_rc/array_down.png'))
        self.popup_widget = PopupWidget(self)
        self.changeElement = self.popup_widget.change_element

        hbox = QHBoxLayout(self)
        hbox.setContentsMargins(1, 1, 1, 1)
        hbox.setSpacing(0)
        hbox.addWidget(self.line)
        hbox.addWidget(self.btn)
        self.btn.clicked.connect(self.show_popup)
        self.popup_widget.state_changed.connect(lambda x: self.line.setText(x))

    def addElements(self, elements):  # {data : is_check}
        self.popup_widget.view.model.load(elements)
        self.line.setText(self.popup_widget.view.model.get())

    def show_popup(self):
        if not self.popup_widget.isVisible() and self.popup_widget.time_for_open.end() >= 0.4:  # если сейчас окно закрыто и с момента закрытия прошло 0.4 секунды
            self.popup_widget.setGeometry(100, 200, 100, 100)
            self.popup_widget.resize(self.width(), min(int(27.6 * 6) + 1, 29 * len(
                self.popup_widget.view.model._data)))  # 29 - ширина одного элемента (+-)
            self.popup_widget.move(self.mapToGlobal(QPoint(0, self.height())))
            self.setFont(self.font())
            self.popup_widget.show()


class DataBase():
    def __init__(self):
        self.con = sqlite3.connect('Timetable.db')
        cur = self.con.cursor()

        cur.execute("""PRAGMA foreign_keys = ON;""")

        cur.execute("""CREATE TABLE IF NOT EXISTS Teacher
        (
          fullName TEXT,
          id INTEGER PRIMARY KEY AUTOINCREMENT
        );
        """)

        cur.execute("""CREATE TABLE IF NOT EXISTS Lesson
        (
          title TEXT,
          hour INT,
          audienceNumber TEXT,
          teacherId INT,
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          FOREIGN KEY (teacherId) REFERENCES Teacher(id) ON DELETE SET NULL ON UPDATE CASCADE
        );
        """)

        cur.execute("""CREATE TABLE IF NOT EXISTS GroupTimetable
        (
          weekNumber INT,
          dayOfWeek INT,
          lessonNumber INT,
          lessonId INT,
          FOREIGN KEY (lessonId) REFERENCES Lesson(id) ON DELETE SET NULL ON UPDATE CASCADE
        );""")

        cur.execute("""CREATE TABLE IF NOT EXISTS Groups
        (
          title TEXT,
          courseNumber INT,
          PRIMARY KEY (title)
        );""")

        cur.execute("""CREATE TABLE IF NOT EXISTS TeacherAttendance
        (
          dayOfWeek INT,
          lessonNumber INT,
          teacherId INT,
          FOREIGN KEY (teacherId) REFERENCES Teacher(id) ON DELETE CASCADE ON UPDATE CASCADE
        );
        """)

        cur.execute("""CREATE TABLE IF NOT EXISTS GroupLesson
        (
          lessonId INT,
          groupTitle TEXT,
          FOREIGN KEY (lessonId) REFERENCES Lesson(id) ON DELETE CASCADE ON UPDATE CASCADE,
          FOREIGN KEY (groupTitle) REFERENCES Groups(title) ON DELETE CASCADE ON UPDATE CASCADE
        );""")

        cur.close()

    def addTeacher(self, teacher):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO Teacher VALUES (?,null)""", (teacher,))
        self.con.commit()

        cur.execute(f"""SELECT id FROM Teacher WHERE fullName = '{teacher}'""")
        for id in cur.fetchall():
            pass
        self.addTeacherTimeWork(id[0])
        cur.close()
        return id[0]

    def addLesson(self, teacher, hours, lesson):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO Lesson VALUES (?,?,?,?,null)""", (lesson, hours, '', teacher))
        self.con.commit()

        cur.execute(f"""SELECT id FROM Lesson WHERE teacherId = '{teacher}'""")
        for id in cur.fetchall():
            pass
        cur.close()
        return id[0]

    def addGroup(self, group, courseNumber):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO Groups VALUES (?,?)""", (group, courseNumber))
        self.con.commit()
        cur.close()

    def get_error(self):
        """чтобы проверить сможет ли учитель вести столько пар в неделю"""
        cur = self.con.cursor()
        cur.execute(f"""SELECT Teacher.fullName, Teacher.id, GroupLesson.groupTitle, Lesson.hour FROM GroupLesson, Lesson, Teacher WHERE Teacher.id = Lesson.teacherId and GroupLesson.lessonId = Lesson.id GROUP BY Lesson.id""")
        teacher_num_lesson = {}
        group_week, error = self.get_group_week()
        for fullName, teacher_id, group, hour in cur.fetchall():
            if (fullName,teacher_id) not in teacher_num_lesson:
                teacher_num_lesson[(fullName,teacher_id)] = 0
            teacher_num_lesson[(fullName,teacher_id)] += hour // group_week[group]
            if fullName == 'Тоболина Т.А.':
                print(teacher_id, hour, group, group_week[group])
        for teacher, num_lesson in teacher_num_lesson.items():
            fullName, teacher_id = teacher
            teacher_num_work_lesson = sum(map(lambda x: x.count(-1), self.getTeacherTimeWork(teacher_id))) * 2     # кол-во рабочих пар у учителя за 2 недели
            if num_lesson > teacher_num_work_lesson:
                error.append(f'У учителя {fullName} слишком много пар - {num_lesson} из {teacher_num_work_lesson} возможных')
        return error

    def getTeacher(self):
        cur = self.con.cursor()
        cur.execute("""SELECT fullName, id FROM Teacher""")
        data = {}
        for rec in cur.fetchall():
            data[rec[1]] = rec[0]
        cur.close()
        return data

    def getTitleTable(self, titleTable, titlePKColumn):
        self.con.row_factory = sqlite3.Row
        cursor = self.con.execute(f'select * from {titleTable}')
        row = cursor.fetchone()
        names = row.keys()
        return names.index(titlePKColumn)

    def getTitleGroup(self):
        cur = self.con.cursor()
        cur.execute("""SELECT title FROM Groups""")
        data = [d[0] for d in cur.fetchall()]
        cur.close()
        return data

    def delRow(self, id, titleTable, titlePKColumn):
        cur = self.con.cursor()
        cur.execute(f"""DELETE FROM {titleTable} WHERE {titlePKColumn} = '{id}'""")
        self.con.commit()
        cur.close()

    def addGroupLesson(self, id_lesson, title_group):
        cur = self.con.cursor()
        cur.execute(f"""INSERT INTO GroupLesson VALUES (?, ?)""", (id_lesson, title_group))
        self.con.commit()
        cur.close()

    def delGroupLesson(self, id_lesson, title_group):
        cur = self.con.cursor()
        cur.execute(f"""DELETE FROM GroupLesson WHERE lessonId = '{id_lesson}' AND groupTitle = '{title_group}'""")
        self.con.commit()
        cur.close()

    def getGroupLesson(self, idLesson):
        cur = self.con.cursor()
        cur.execute(f"""SELECT groupTitle FROM GroupLesson WHERE lessonId = '{idLesson}' """)
        data = [d[0] for d in cur.fetchall()]
        cur.close()
        return data

    def addTeacherTimeWork(self, id_teacher):
        cur = self.con.cursor()
        for day in range(6):
            for lesson in range(4):
                cur.execute(f"""INSERT INTO TeacherAttendance VALUES (?,?,?)""", (day, lesson, id_teacher))
        self.con.commit()
        cur.close()

    def getTeacherTimeWork(self, id_teacher):
        '''рабочие дни'''
        cur = self.con.cursor()
        timeWork = [[False, False, False, False] for _ in range(6)]
        cur.execute(f"""SELECT dayOfWeek, lessonNumber FROM TeacherAttendance WHERE teacherId = {id_teacher}""")
        for day, lesson in cur.fetchall():
            timeWork[day][lesson] = -1
        return timeWork

    def add_work_day(self, day, les, id):
        """удаление и добавление рабочего дня для учителя"""
        cur = self.con.cursor()
        cur.execute("""INSERT INTO TeacherAttendance VALUES (?,?,?)""", (day, les, id))
        self.con.commit()
        cur.close()

    def delete_work_day(self, id_teacher, day, lesson):
        cur = self.con.cursor()
        cur.execute(f"""DELETE FROM TeacherAttendance WHERE dayOfWeek = {day} and lessonNumber = {lesson} and teacherId = {id_teacher}""")
        self.con.commit()
        cur.close()

    # def get_teachers_for_group_in_day(self):
    #     """для генерации расписания {day:{numLesson:{group:[teachers]}}}"""
    #     cur = self.con.cursor()
    #     cur.execute("""SELECT DISTINCT TeacherAttendance.dayOfWeek, TeacherAttendance.lessonNumber, GroupLesson.groupTitle, TeacherAttendance.teacherId
    #                     FROM GroupLesson, Lesson, TeacherAttendance WHERE TeacherAttendance.teacherId = Lesson.teacherId and GroupLesson.lessonId = Lesson.id""")
    #     data = {}
    #     for day, numLesson, group, teacher in cur.fetchall():
    #         if day not in data.keys():
    #             data[day] = {}
    #         if numLesson not in data[day].keys():
    #             data[day][numLesson] = {}
    #         if group not in data[day][numLesson].keys():
    #             data[day][numLesson][group] = []
    #         data[day][numLesson][group].append(teacher)
    #     return data

    # def get_group_teacher_lesson(self):
    #     """{group:teacher:[[lessonID, quantityLessonForTwoWeek], ...]}"""
    #     cur = self.con.cursor()
    #     cur.execute("""SELECT DISTINCT GroupLesson.groupTitle, Lesson.teacherId, GroupLesson.lessonId, Lesson.hour FROM GroupLesson, Lesson WHERE GroupLesson.lessonId = Lesson.id""")
    #     query = cur.fetchall()
    #     data = {}
    #     group_hours = self.get_hours_group()
    #     for group, teacher, lesson, hour in query:
    #         if group not in data.keys():
    #             data[group] = {}
    #             group_hours[group] //= 36   # получаем кол-во учебных недель
    #         if teacher not in data[group].keys():
    #             data[group][teacher] = []
    #         data[group][teacher].append([lesson, hour//group_hours[group]])
    #     return data, group_hours

    def get_group_week(self):
        """кол-во недель для группы за семестр"""
        cur = self.con.cursor()
        cur.execute("""SELECT GroupLesson.groupTitle, SUM(Lesson.hour) FROM GroupLesson, Lesson WHERE Lesson.id = GroupLesson.lessonId GROUP BY GroupLesson.groupTitle""")
        group_week = {}
        error = []
        for group, hours in cur.fetchall():
            if hours%36 != 0:
                error.append(f'У группы {group} указано неверно кол-во часов - {hours}')
            group_week[group] = hours//36
        cur.close()
        return group_week, error

    def getAllTeacherTimeWork(self):
        """получение всех возможных дней работы учителей"""
        cur = self.con.cursor()
        cur.execute("""SELECT id FROM Teacher""")
        teacherID_TimeWork = {}
        # TW_teacher = self.getTeacherTimeWork(id)
        # for

        for id in cur.fetchall():
            id = id[0]
            teacherID_TimeWork[id] = [self.getTeacherTimeWork(id), self.getTeacherTimeWork(id)]
        cur.close()
        return teacherID_TimeWork

    # def get_group_lesson_before(self):
    #     cur = self.con.cursor()
    #     cur.execute("""SELECT Lesson.hour, lessonId, groupTitle FROM GroupLesson, Lesson WHERE GroupLesson.lessonId = Lesson.id""")
    #     number_free_lessons = 12
    #     group_lesson = {}
    #     group_hours = self.get_hours_group()
    #     # print(group_hours)
    #     for hour, lesson, group in cur.fetchall():
    #         if group not in group_lesson.keys():
    #             group_lesson[group] = [None]*number_free_lessons
    #         group_lesson[group] += [lesson]*(hour//(group_hours[group]//36))
    #         # print([lesson]*(hour//(group_hours[group]//36)))
    #         # print(lesson, hour//(group_hours[group]//36), hour, group, group_hours[group])
    #     return group_lesson

    def get_group_lesson(self):
        """для GA"""
        cur = self.con.cursor()
        cur.execute("""SELECT Lesson.hour, lessonId, groupTitle FROM GroupLesson, Lesson WHERE GroupLesson.lessonId = Lesson.id ORDER BY groupTitle""")
        number_free_lessons = 12
        group_week, error = self.get_group_week()
        lessonsID = {}
        # group_lesson = []
        # index_group = []
        # group_before = None
        # index_lesson = []
        # for hour, lesson, group in cur.fetchall():
        #     if group_before != group:
        #         group_before = group
        #         i = 0 + number_free_lessons
        #         group_lesson.append([i for i in range(number_free_lessons)])
        #         index_group.append(group)
        #         index_lesson[group].append([None] * number_free_lessons)
        #     group_lesson[-1] += [i for i in range(i, i + hour // group_week[group])]
        #     i += hour // group_week[group]
        #     index_lesson[-1] += [lesson] * (hour // group_week[group])
        #     # print([lesson]*(hour//(group_hours[group]//36)))
        #     # print(lesson, hour//(group_hours[group]//36), hour, group, group_hours[group])
        # return group_lesson, lessonsID, index_group
        for hour, lesson, group in cur.fetchall():
            if group not in lessonsID:
                lessonsID[group] = [None] * number_free_lessons
            lessonsID[group] += [lesson] * (hour // group_week[group])
        return lessonsID

    def get_lessonID_teacherID(self):
        """для GA"""
        cur = self.con.cursor()
        cur.execute("""SELECT id, teacherId FROM Lesson""")
        lessonID_teacherID = {}
        for lessonID, teacherID in cur.fetchall():
            lessonID_teacherID[lessonID] = teacherID
        cur.close()
        return lessonID_teacherID

    def insert_time_table(self, timeTable):
        """после генерации расписания"""
        cur = self.con.cursor()
        cur.execute("""DELETE FROM GroupTimetable""")
        # weekNumber dayOfWeek lessonNumber lessonId
        for group_lesson in timeTable:
            for week in range(2):
                for day in range(6):
                    for num_lesson in range(4):
                        num = week*6*4 + day*4 + num_lesson
                        lesson = group_lesson[num]
                        if lesson != None:
                            cur.execute("""INSERT INTO GroupTimetable VALUES (?,?,?,?)""", (week, day, num_lesson, lesson))
        cur.close()
        self.con.commit()

    def getTimeTable(self):
        cur = self.con.cursor()
        cur.execute("""SELECT groupTitle, weekNumber, dayOfWeek, lessonNumber, GroupLesson.lessonId FROM GroupTimetable, GroupLesson WHERE GroupLesson.lessonId = GroupTimetable.lessonId""")
        TimeTable = {}
        data = cur.fetchall()
        for group, week, day, num_lesson, lesson_id in data:
            if group not in TimeTable.keys():
                TimeTable[group] = {}
            if week not in TimeTable[group].keys():
                TimeTable[group][week] = {}
            if day not in TimeTable[group][week].keys():
                TimeTable[group][week][day] = {}
            TimeTable[group][week][day][num_lesson] = lesson_id
        return TimeTable

    def getGroup(self):
        cur = self.con.cursor()
        cur.execute("""SELECT title FROM Groups""")
        groups = []
        data = cur.fetchall()
        for group in data:
            groups.append(group)
        cur.close()
        return groups

    def get_LessonID_Teacher(self):
        cur = self.con.cursor()
        cur.execute("""SELECT Lesson.id, title, fullName, audienceNumber FROM Lesson, Teacher WHERE Lesson.TeacherId = Teacher.id""")
        lessonID_Teacher = {}
        data = cur.fetchall()
        for lessonID, lesson, teacher, audienceNumber in data:
            lessonID_Teacher[lessonID] = [lesson, teacher, audienceNumber]
        cur.close()
        return lessonID_Teacher

    def getTeacherTimeWorkLesson(self, id):
        """Получаем расписание учителя"""
        cur = self.con.cursor()
        cur.execute(f"""SELECT weekNumber, dayOfWeek, lessonNumber, Lesson.title, GroupLesson.groupTitle  FROM GroupTimetable, Lesson, GroupLesson  WHERE GroupTimetable.lessonId = Lesson.id and Lesson.teacherId = {id} and GroupLesson.lessonId = Lesson.id""")
        timeTableTeacher = {}
        for week, day, num_lesson, title_lesson, title_group in cur.fetchall():
            if week not in timeTableTeacher.keys():
                timeTableTeacher[week] = {}
            if day not in timeTableTeacher[week].keys():
                timeTableTeacher[week][day] = {}
            if num_lesson not in timeTableTeacher[week][day].keys():
                timeTableTeacher[week][day][num_lesson] = [title_lesson, []]
            timeTableTeacher[week][day][num_lesson][1].append(title_group)
        cur.close()
        return timeTableTeacher

    def get_paried_lesson(self):
        """для GA получаем спаренные занятия"""
        cur = self.con.cursor()
        data = cur.execute("""SELECT lessonId, GroupLesson.groupTitle, Lesson.hour FROM GroupLesson, Lesson WHERE Lesson.id = lessonId group by 1 having count(lessonId) > 1 """)
        # pariedLesson_groups = {}
        lessons_par = {}
        group_week, error = self.get_group_week()
        for lessonID, group, hour in data:
            lessons_par[lessonID] = {'всего занятий':hour//group_week[group], 'группы':[]}
        # return lessons_par
        for lessonID in lessons_par:
            for group, in cur.execute(f"""SELECT groupTitle FROM GroupLesson WHERE lessonId = {lessonID}""").fetchall():
                lessons_par[lessonID]['группы'].append(group)
        return lessons_par
        # return pariedLesson_groups

class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.DB = DataBase()
        self.ui = PyQt5.uic.loadUi("main.ui", self)
        setStyle(self.ui)
        self.ui.show()
        # self.setFont(QFont("Comics Sans MS", 10))
        # self.clickBtnAdd = False
        # self.xclx = ParsingXLSX(self.DB)
        self.teacherTable = tableWithEditing(['Учителя', 'id', '', ''], "Teacher", 'id', self.DB, extra_button=True)
        self.ui.horizontalLayout.addWidget(self.teacherTable)
        self.groupTable = tableWithEditing(['Группы', 'Курс', ''], "Groups", 'title', self.DB)
        self.ui.horizontalLayout_2.addWidget(self.groupTable)
        self.lessonTable = tableWithEditing(['Занятие', 'Часы', 'Кабинет', 'Учитель', 'id', 'Группа', ''], "Lesson", 'id',
                                  self.DB, self.ui)
        self.ui.horizontalLayout_3.addWidget(self.lessonTable)
        # timeTable = TimeTable(self.DB, self.ui)
        # layout = QtWidgets.QGridLayout()
        # TimeTableTab = TimeTable(self.DB)
        # self.ui.tabWidget.addTab(QWidget(), "Расписание")
        # self.ui.tabWidget.
        # QtWidgets.QTabWidget.addTab()
        # GenerateTimeTable(self.DB)
        # GeneticGenerationTimeTable(self.DB)
        menuBar = self.ui.menuBar()
        self.fileMenu = QMenu("&Файл", self)
        menuBar.addMenu(self.fileMenu)
        self.actionOpen = QAction("Загрузить данные из excel", self)
        self.actionSave = QAction("Сохранить расписание в excel", self)
        self.fileMenu.addAction(self.actionOpen)
        self.fileMenu.addAction(self.actionSave)
        self.actionOpen.triggered.connect(self.importData)
        self.actionSave.triggered.connect(self.exportTimeTable)
        self.setWindowTitle('Расписание')
        # QPushButton.style()
        self.ui.btn_next_TimeTable.hide()
        self.ui.progressBar.hide()
        self.ui.lcdNumber.hide()
        self.ui.btn_generate_TimeTable.clicked.connect(self.generate_TimeTable)
        # print(set([1,2,3])-set([2,3,4]))
        # print('er')
        # print(self.DB.get_paried_lesson())
        # self.time = 100
        # self.timer = QTimer(self)
        # self.timer.timeout.connect(self.showTime)
        # self.timer.setInterval(10)  # 1 sec
        # self.showTime()
    # def any(self):
    #     from openpyxl import Workbook
    #     # workbook = load_workbook('C:\\Users\\aleks\\OneDrive\\Рабочий стол\\test123.xlsx')
    #     workbook = Workbook()
    #     worksheet = workbook.active
    #     # worksheet = workbook['Sheet1']
    #     print(worksheet)
    #     worksheet.cell(row=1, column=2).value = '12345'
    #     print(worksheet.cell(row=1, column=2).value)
    #     worksheet.merge_cells(start_row=5, end_row=6, start_column=2, end_column=2)
    #     workbook.save('C:\\Users\\aleks\\OneDrive\\Рабочий стол\\test123.xlsx')
    #     workbook.close()

    def generate_TimeTable(self):
        print('create timetable')
        errors = self.DB.get_error()
        if errors != []:
            msg = QMessageBox.warning(
                self,
                "Success!",
                '\n'.join(errors)
            )
            return

        self.btn_generate_TimeTable.setEnabled(False)
        self.tabWidget.setEnabled(False)
        self.fileMenu.setEnabled(False)
        self.ui.progressBar.show()
        self.ui.lcdNumber.show()
        self.ui.progressBar.setEnabled(True)
        self.ui.lcdNumber.setEnabled(True)
        self.ui.lcdNumber.setNumDigits(8)

        self.time = Timer()
        self.timer = QTimer(self)
        self.timer.setInterval(1000)  # 1 sec
        self.timer.timeout.connect(self.showTime)
        self.timer.start()

        self.generate_TT = GA(self.DB.get_group_lesson(), self.DB.getAllTeacherTimeWork(), self.DB.get_lessonID_teacherID(), self.DB.get_paried_lesson())
        self.generate_TT.progress.connect(lambda x: self.setStatusProgressBar(x))
        self.generate_TT.result.connect(lambda TT: self.end_GA(TT))
        self.generate_TT.start()

    def end_GA(self, TimeTable_for_save):
        self.DB.insert_time_table(TimeTable_for_save)
        self.setStatusProgressBar(100)
        self.timer.stop()
        self.btn_generate_TimeTable.setEnabled(True)
        self.fileMenu.setEnabled(True)
        self.tabWidget.setEnabled(True)
        table = self.ui.TimeTable
        for row in range(48):
            try:
                print(row, table.item(row, 1).text().replace('\n', ' '), end='\t')
            except Exception:
                pass
            try:
                print(table.item(row, 2).text().replace('\n', ' '))
            except Exception:
                pass
        time_table = TimeTable(self.DB, self.ui)
        time_table.load_TimeTable()
        for row in range(48):
            try:
                print(row, table.item(row, 1).text().replace('\n', ' '), end='\t')
            except Exception:
                pass
            try:
                print(table.item(row, 2).text().replace('\n', ' '))
            except Exception:
                pass

    def showTime(self):
        sec = int(self.time.end())%60
        if sec < 10:
            sec = '0' + str(sec)
        else:
            sec = str(sec)
        self.ui.lcdNumber.display(str(int(self.time.end()//60)) + ':' + sec)

    def setStatusProgressBar(self, progress):
        self.ui.progressBar.setValue(progress)

    def update_Tables(self):
        self.teacherTable.setTableData()
        self.groupTable.setTableData()
        self.lessonTable.setTableData()

    def exportTimeTable(self):
        from openpyxl import Workbook
        self.table = self.ui.TimeTable
        fileName, ok = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл",
            ".",
            "All Files(*.xlsx)"
        )
        print(fileName)
        if not fileName:
            return

        workbook = Workbook()
        worksheet = workbook.active
        TimeTable = self.DB.getTimeTable()
        print(TimeTable)
        groups = self.DB.getTitleGroup()
        LessonID_teacher = self.DB.get_LessonID_Teacher()
        print(groups)
        print(LessonID_teacher)
        days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
        for num_day, day in enumerate(days):
            worksheet.merge_cells(start_row=num_day*8+2, end_row=num_day*8+9, start_column=1, end_column=1)
            worksheet.cell(row=num_day*8+2, column=1).value = day

        for num_group, group in enumerate(groups):
            worksheet.cell(row=1, column=num_group+2).value = group

        shift_col = 2
        shift_row = 2
        for num_group in range(len(groups)):
            for week in TimeTable[groups[num_group]]:
                for day in TimeTable[groups[num_group]][week]:
                    for num_lesson in TimeTable[groups[num_group]][week][day]:
                        lessonID = TimeTable[groups[num_group]][week][day][num_lesson]
                        lesson, teacher, audienceNumber = LessonID_teacher[lessonID]
                        data_for_cell = lesson + '\n' + teacher
                        row = day*2*4 + num_lesson*2 + week + shift_row # плюс 1 тк отсчет в xclx с 1 и заголовки
                        col = num_group+shift_col
                        # worksheet.write(day*2*4 + num_lesson*2 + week, num_group+shift_group, data_for_cell)
                        worksheet.cell(row=row, column=col).value = data_for_cell
                        if week == 1:
                            text = worksheet.cell(row=row-1, column=col).value
                            if text == data_for_cell:
                                worksheet.merge_cells(start_row=row-1, end_row=row, start_column=col, end_column=col)

        # если одинаковые ячейки, то оъединяем их
        for col in range(shift_col, len(groups)+shift_col):
            for row in range(shift_row, shift_row+(2*4*6), 2):
                if worksheet.cell(row=row, column=col).value == worksheet.cell(row=row+1, column=col).value:
                    self.table.setSpan(row, col, 2, 1)
                    worksheet.merge_cells(start_row=row, end_row=row+1, start_column=col, end_column=col)

        workbook.save(fileName)
        workbook.close()
        msg = QMessageBox.information(
            self,
            "Success!",
            f"Данные сохранены в файле: \n{fileName}"
        )

    def importData(self):
        win = ParsingXLSX(self.DB, parent=self)
        win.setWindowTitle('Импорт')
        win.exec()


class ParsingXLSX(QDialog):
    def __init__(self, DB, parent = None):
        super().__init__(parent)
        self.parent = parent
        self.ui = PyQt5.uic.loadUi('import.ui', self)
        setStyle(self.ui)
        self.ui.btn_import_data.clicked.connect(self.select_file)
        self.DB = DB
        self.ui.show()
        self.setWindowIcon(QIcon('img\import.jpg'))

    def select_file(self):
        fileName, ok = QFileDialog.getOpenFileName(
            self,
            "Сохранить файл",
            ".",
            "All Files(*.xlsx)"
        )
        self.file = openpyxl.load_workbook(fileName)
        self.teacher_group_lesson = {}
        self.get_teacher_from_xlsx(start_row=self.num_start_row.value(), col_lesson=self.num_col_lesson.value(), col_techaer=self.num_col_teacher.value(), col_hours_1=self.num_col_hours.value(), col_hours_2=self.num_col_hours_extra.value())
        self.save_in_DB()
        self.parent.update_Tables()
        msg = QMessageBox.information(
            self,
            "Success!",
            f"Данные импортированы из файла: \n{fileName}"
        )

    def get_teacher_from_xlsx(self, start_row=9, col_lesson = 2, col_techaer=3, col_hours_1=9, col_hours_2=10):
        for group in self.file.sheetnames:
            now_row = start_row
            work_list = self.file[group]
            while work_list.cell(row = now_row, column = col_techaer).value != None:
                hours = 0
                hours_1 = work_list.cell(row = now_row, column = col_hours_1).value
                hours_2 = work_list.cell(row = now_row, column = col_hours_2).value
                hours += int(hours_1) if hours_1 else 0
                hours += int(hours_2) if hours_2 else 0
                if hours != 0:
                    lesson = work_list.cell(row = now_row, column = col_lesson).value
                    teacher = work_list.cell(row = now_row, column = col_techaer).value
                    try:
                        self.teacher_group_lesson[teacher][group].append([hours, lesson])
                    except KeyError:    # если у учителя еще нет такой группы
                        try:
                            self.teacher_group_lesson[teacher][group] = [[hours, lesson]]
                        except KeyError:    # если еще не было такого учителя
                            self.teacher_group_lesson[teacher] = {}
                            self.teacher_group_lesson[teacher][group] = [[hours, lesson]]
                now_row += 1
        print(self.teacher_group_lesson)

    def save_in_DB(self):
        is_group = {}
        for teacher in self.teacher_group_lesson.keys():
            id = self.DB.addTeacher(teacher)
            # print(id, teacher)
            # teacher_id[teacher] = id
            for group in self.teacher_group_lesson[teacher]:
                try:
                    if is_group[group]:
                        pass
                except:
                    is_group[group] = True
                    for course in group:
                        if course.isdigit():
                            break
                    self.DB.addGroup(group, int(course))
                for hours, lesson in self.teacher_group_lesson[teacher][group]:
                    id_lesson = self.DB.addLesson(id, hours, lesson)
                    self.DB.addGroupLesson(id_lesson, group)


class tableWithEditing(QtWidgets.QWidget):
    def __init__(self, headTable, titleTable, titlePKColumn, DB, ui=False, extra_button=False, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.layout = QtWidgets.QGridLayout()
        self.titlePKColumn = titlePKColumn  # для удаления столбца
        self.setLayout(self.layout)
        self.titleTable = titleTable
        self.tableData = QtWidgets.QTableView()
        self.tableData.setSortingEnabled(True)
        self.layout.addWidget(self.tableData)

        con = QSqlDatabase.addDatabase("QSQLITE")
        con.setDatabaseName('Timetable.db')
        con.open()

        # model = QSqlTableModel()
        # model.setTable(self.titleTable)
        # model.select()
        # self.model = model
        # self.tableData.setModel()

        self.headTable = headTable
        self.DB = DB
        self.clickBtnAdd = False
        self.extra_button = extra_button

        self.setBtnAdd()
        self.setTableData()
        # self.setWidgets(title=True)
        self.model.dataChanged.connect(self.changeTable)
        # self.changeTable()
        if ui:
            ui.tabWidget.currentChanged.connect(self.changeTab)
            self.tabWidget = ui.tabWidget
            self.ui = ui
        # self.tableData.setTextElideMode(Qt.ElideNone)
        self.tableData.horizontalHeader().setStretchLastSection(True)
        self.tableData.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.tableData.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.tableData.horizontalHeader().setSectionResizeMode(self.model.columnCount()-2, 1)
        # self.tableData.horizontalHeader().setSectionResizeMode(self.model.columnCount()-1, 1)

    def setWidgets(self, title=False):
        """
        после .select() нужно вставить виджеты и скрыть колонку id
        :param title: при первом заполнении таблицы так же нужно вставить заголовки
        :return:
        """
        self.model.insertColumn(self.model.columnCount())  # для кнопок удаления
        if self.titleTable == "Teacher" or self.titleTable == "Lesson":
            self.tableData.setColumnHidden(self.model.columnCount() - 2,
                                           True)  # спрятать id (предпоследняя колонна) для Teacher or Lesson
        if self.titleTable == "Groups":
            self.setSpinBox()
        elif self.titleTable == "Lesson":
            self.setSpinBox(1, 99999)  # добавляем spinbox
            self.setComboBox(3, self.DB.getTeacher())
            self.setCheckableComboBox(self.DB.getTitleGroup())

        if self.extra_button:
            self.setExtraButton()  # используем столбец для кнопки удаления
            self.model.insertColumn(self.model.columnCount())  # и добавляем новый столбец для кнопки удаления

        self.setBtnDel()
        if title:
            for i, head in enumerate(self.headTable):
                self.model.setHeaderData(i, Qt.Horizontal, head)
        self.changeTable()
        self.setStyleSheet("""
        QPushButton { text-align: center;}
        QLabel {background-color: rgb(255,128,0);}
        QWidget {font-size:16px; font-family:Comic Sans MS;}
        """)

    def setExtraButton(self):
        row = self.model.rowCount()  # кол-во строк
        self.extraBtn_row = {}
        for y in range(row):
            extra_btn = QPushButton()
            size = 37
            extra_btn.setMinimumSize(size, size)
            extra_btn.setMaximumSize(size, size)  # x, y
            icon = QIcon("img/teach_negative.png")
            extra_btn.setIconSize(QtCore.QSize(size - 7, size - 7))
            extra_btn.setIcon(icon)
            cell_widget = QWidget()
            lay_out = QHBoxLayout(cell_widget)
            lay_out.addWidget(extra_btn)
            lay_out.setAlignment(QtCore.Qt.AlignCenter)
            lay_out.setContentsMargins(0, 0, 0, 0)  # Рисуем границы
            cell_widget.setLayout(lay_out)
            self.extraBtn_row[extra_btn] = y
            self.tableData.setIndexWidget(self.model.index(y, self.model.columnCount() - 1), cell_widget)
            extra_btn.clicked.connect(lambda: self.openWorkingTime())

    def openWorkingTime(self):
        btn = self.sender()
        row = self.extraBtn_row[btn]
        # print(row, id_Teacher := self.model.index(row, 0).data(), id_Teacher := self.model.index(row, 1).data())
        # id_Teacher = self.model.index(0, row).data()
        # id_Teacher = self.model.index(1, row).data()
        id_Teacher = self.model.index(row, 1).data()
        win = WorkingTime(name=self.model.index(row, 0).data(), id=id_Teacher, DB=self.DB)
        setStyle(win, selectCell=False)
        win.show()
        win.exec()

    def setFromCheckableComboBox(self, el, is_check, comboBox):
        row = self.сheckableСomboBox_row[comboBox]
        id_lesson = self.model.index(row, 4).data()
        if is_check:
            self.DB.addGroupLesson(id_lesson, el)
        else:
            self.DB.delGroupLesson(id_lesson, el)

    def setCheckableComboBox(self, data):
        self.model.insertColumn(self.model.columnCount())
        column = self.model.columnCount() - 2
        self.сheckableСomboBox_row = {}
        for row in range(self.model.rowCount()):
            data_check = {d: False for d in data}
            d_cheked = self.DB.getGroupLesson(self.model.index(row, 4).data())
            for d in d_cheked:
                data_check[d] = True
            сheckableСomboBox = MyCheckableComboBox()
            сheckableСomboBox.changeElement.connect(
                lambda el, is_check, comboBox: self.setFromCheckableComboBox(el, is_check, comboBox))
            сheckableСomboBox.addElements(data_check)
            # сheckableСomboBox.addItems(data)
            # сheckableСomboBox.lineEdit().setText('Выберите что-то')
            # nowId = self.model.index(row, column).data()  # берем из таблицы внешней ключ
            # if nowId == '' and data:
            # self.updateRow(row, column, int(сheckableСomboBox.currentText().split('|')[1]))
            # comboBox.setStyleSheet("""background-color:rgb(255,128,138);""")
            # elif nowId != '':
            #     сheckableСomboBox.setCurrentText(data[nowId]+' | '+str(nowId))
            self.tableData.setIndexWidget(self.model.index(row, column), сheckableСomboBox)
            self.сheckableСomboBox_row[сheckableСomboBox] = row
            # сheckableСomboBox.currentIndexChanged.connect(self.setFromComboBox)
            # сheckableСomboBox.popup_widget.view.model._checklist[0] = True
        # print(self.сheckableСomboBox_row)
        # print(self.checkableComboBox_row)     # в чем отличие от прошлой строчки??????

    def addRowTable(self):
        """добавление строки в таблицу при нажатии на кнопку"""
        # print('123')
        self.clickBtnAdd = True
        self.model.submitAll()  # ячейка при добавлении новой может еще редактироваться пользователем, поэтому надо сначала принять введенное (при редактировании нельзя добавить)
        record = self.model.record()
        if self.titleTable == "Teacher":
            record.remove(record.indexOf("id"))
            record.setValue("fullName", "")
        elif self.titleTable == "Groups":
            record.setValue("title", "")
            record.setValue("courseNumber", 1)
        elif self.titleTable == "Lesson":
            record.remove(record.indexOf("id"))
            record.setValue("title", "")
            record.setValue("hour", 1)
            record.setValue("audienceNumber", 1)
            record.setValue("teacherId", "")
        self.model.insertRecord(-1, record)

        self.model.select()
        self.setWidgets()
        if self.titleTable == "Teacher":
            id_teacher = self.model.index(self.model.rowCount() - 1, 1).data()
            self.DB.addTeacherTimeWork(id_teacher)

        self.tableData.scrollToBottom()  # пролистываем вниз
        self.clickBtnAdd = False

    def setBtnAdd(self):
        """добавляем кнопку добавления новой строки"""
        btnAdd = QtWidgets.QPushButton()
        self.layout.addWidget(btnAdd)
        btnAdd.clicked.connect(self.addRowTable)
        btnAdd.setStyleSheet("""background-color: rgb(255,255,255); text-align: center;""")
        size = 30
        btnAdd.setMinimumSize(size, size)
        btnAdd.setIconSize(QtCore.QSize(size - 7, size - 7))

        icon = QIcon("img/add.png")
        btnAdd.setIcon(icon)

    def delRowTable(self):
        """удаление строки из таблицы и бд при нажатии на кнопку"""
        btnDel = self.sender()
        row = self.tableData.indexAt(btnDel.pos()).row()
        column = self.DB.getTitleTable(self.titleTable, self.titlePKColumn)  # по названию возвращаем номер столбца
        dataId = self.model.index(row, column).data()
        self.DB.delRow(dataId, self.titleTable, self.titlePKColumn)  # из-за каскадного удаления
        self.model.submitAll()
        self.model.select()
        self.setWidgets()

    def setBtnDel(self):
        """установка кнопок удаления в таблицу"""
        row = self.model.rowCount()  # кол-во строк
        for y in range(row):
            btnDel = QPushButton()
            size = 37
            btnDel.setMinimumSize(size, size)
            btnDel.setMaximumSize(size, size)  # x, y
            icon = QIcon("img/del.png")
            btnDel.setIcon(icon)
            btnDel.setIconSize(QtCore.QSize(size - 17, size - 17))
            # btnDel.setStyleSheet("""padding: 0px 230px 0px;""")
            self.tableData.setIndexWidget(self.model.index(y, self.model.columnCount() - 1), btnDel)
            btnDel.clicked.connect(lambda: self.delRowTable())

    def setSpinBox(self, column=1, max=5):
        @pyqtSlot()
        def setFromSpinBox(tableData):  # вызывается при установке значения из spinBox в БД
            self.model.submitAll()
            spinBox = self.sender()
            rowSpinBox = tableData.indexAt(spinBox.pos()).row()
            columnSpinBox = tableData.indexAt(spinBox.pos()).column()
            numForSet = spinBox.value()  # вставляем число в бд
            record = self.model.record()
            for col in range(self.model.columnCount() - 1):
                record.setValue(col, self.model.index(rowSpinBox, col).data())
                if col == columnSpinBox:
                    record.setValue(1, numForSet)
            self.model.updateRowInTable(rowSpinBox, record)

        for row in range(self.model.rowCount()):  # кол-во строк
            spinBox = QSpinBox()
            spinBox.wheelEvent = lambda event: None
            spinBox.setMaximum(max)
            spinBox.setMinimum(1)
            spinBox.setValue(self.model.index(row, column).data())  # вставляем в spinbox данные из ячейки
            self.tableData.setIndexWidget(self.model.index(row, column), spinBox)
            spinBox.editingFinished.connect(lambda: setFromSpinBox(self.tableData))

    def setFromComboBox(self):
        '''при изменении значении в combobox вставляем это значение в БД'''
        # print('комбобокс изменен')
        comboBox = self.sender()
        comboBox.setStyleSheet("""""")
        rowComboBox = self.tableData.indexAt(comboBox.pos()).row()
        columnComboBox = self.tableData.indexAt(comboBox.pos()).column()
        id = int(comboBox.currentText().split('|')[1])
        self.updateRow(rowComboBox, columnComboBox, id)

    def setComboBox(self, column, data):
        """
        вставка QCombobox в таблицу
        :param column: колонка куда вставляется QComboBox
        :param data: данные для вставки {id:data}
        """
        for row in range(self.model.rowCount()):
            comboBox = QComboBox()
            comboBox.addItems([v + ' | ' + str(k) for k, v in data.items()])
            nowId = self.model.index(row, column).data()  # берем из таблицы внешней ключ
            if nowId == '' and data:
                self.updateRow(row, column, int(comboBox.currentText().split('|')[1]))
                comboBox.setStyleSheet("""background-color:rgb(255,128,138);""")
            elif nowId != '' and data:
                comboBox.setCurrentText(data[nowId] + ' | ' + str(nowId))
            self.tableData.setIndexWidget(self.model.index(row, column), comboBox)
            comboBox.currentIndexChanged.connect(self.setFromComboBox)

            comboBox.setEditable(True)  # для поиска в comboBox
            comboBox.setInsertPolicy(QtWidgets.QComboBox.NoInsert)
            # change completion mode of the default completer from InlineCompletion to PopupCompletion
            comboBox.completer().setCompletionMode(QtWidgets.QCompleter.PopupCompletion)

    def updateRow(self, row, column, id):
        """
        обнавляем строку в таблице (при добавлении значения в бд из виджета)
        :param row:
        :param id:
        :return:
        """
        record = self.model.record()
        for col in range(self.model.columnCount() - 1):
            print(self.model.index(row, col).data())
            if col == column:
                record.setValue(col, id)
            else:
                record.setValue(col, self.model.index(row, col).data())
        self.model.updateRowInTable(row, record)

    def changeTable(self):
        """вызывается при изменении ячейки"""
        # print('ячейка изменена')
        # self.tableData.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        # self.tableData.horizontalHeader().setSectionResizeMode(self.model.columnCount()-2, 0)
        # # self.tableData.horizontalHeader().setStretchLastSection(True)
        # self.tableData.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.tableData.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        if not self.clickBtnAdd:  # я не знаю почему, но если вызывать submitAll при добавлении строки, то программа вылетает, но submitAll нужно вызывать, чтобы сразу же сохранить изменения в ячейки таблицы, а не после нажатия на enter
            self.model.submitAll()

    def changeTab(self):
        """
        обновляется таблица для обновления данных в comboBox
        вызывается при переходе на вкладку 'Занятия'
        :return:
        """
        print('вкладка изменена на', self.tabWidget.currentIndex())
        # if self.titleTable == "Teacher" and self.ui.tabWidget.currentIndex() == 0:
        # elif self.titleTable == "Groups" and self.ui.tabWidget.currentIndex() == :
        if self.titleTable == "Lesson" and self.tabWidget.currentIndex() == 2:
            self.model.select()
            self.setWidgets()
        if self.tabWidget.currentIndex() == 3:
            print('расписание')
            time_table = TimeTable(self.DB, self.ui)
            time_table.load_TimeTable()

    def setTableData(self):
        """установка данных в таблицу из БД """
        con = QSqlDatabase.addDatabase("QSQLITE")
        con.setDatabaseName('Timetable.db')
        con.open()

        model = QSqlTableModel()
        model.setTable(self.titleTable)
        model.select()

        self.tableData.setModel(model)
        # self.tableData.verticalHeader().hide()   # спрятать цифры сбоку

        # self.setWidgets(self.model)
        # self.model.insertColumn(self.model.columnCount())    # для удаления
        # if self.titleTable == "Teacher" or self.titleTable == "Lesson":
        #     self.tableData.setColumnHidden(self.model.columnCount()-2, True)  # спрятать id (предпоследняя колонна) для Teacher or Lesson
        # # if self.titleTable == "Lesson":
        # #     self.model.insertColumn(self.model.columnCount()-1)    # для групп

        # for i, head in enumerate(self.headTable):
        #     self.model.setHeaderData(i, Qt.Horizontal, head)
        self.model = model
        self.setWidgets(title=True)
        # return model


class VerticalTextDelegate(QStyledItemDelegate):
    """вертикальный текст"""
    def __init__(self, parent):
        super(VerticalTextDelegate, self).__init__()

    def paint(self, painter, option, index):
        optionCopy = QStyleOptionViewItem(option)
        rectCenter = QtCore.QPointF(QtCore.QRectF(option.rect).center())
        painter.save()
        painter.translate(rectCenter.x(), rectCenter.y())
        painter.rotate(-90.0)
        painter.translate(-rectCenter.x(), -rectCenter.y())
        optionCopy.rect = painter.worldTransform().mapRect(option.rect)
        super(VerticalTextDelegate, self).paint(painter, optionCopy, index)

        painter.restore()

    # def sizeHint(self, option, index):
    #     val = QtGui.QSize(self.sizeHint(option, index))
    #     return QtGui.QSize(val.height(), val.width())


class TimeTable(QWidget):
    """отображение основного расписания"""
    def __init__(self, DB, ui):
        super().__init__()
        self.DB = DB
        self.table = ui.TimeTable
        self.ui = ui

    def load_TimeTable(self):
        print('load TT')
        TimeTable = self.DB.getTimeTable()
        print(TimeTable)
        groups = self.DB.getTitleGroup()
        LessonID_teacher = self.DB.get_LessonID_Teacher()
        print(groups)
        print(LessonID_teacher)

        numrows = 6*8
        numcols = len(groups)+1  # 3 columns in your example
        self.table.clear()
        self.table.setColumnCount(0)
        self.table.setRowCount(0)
        self.table.setColumnCount(numcols)
        self.table.setRowCount(numrows)
        self.table.setHorizontalHeaderLabels(['День']+groups)


        self.table.setItemDelegateForColumn(0, VerticalTextDelegate(self))  # дни пишем вертикально
        # self.table.setItemDelegateForColumn(1, VerticalTextDelegate(self))  # дни пишем вертикально

        # delegate = AlignDelegate(self.table)
        # self.table.setItemDelegateForColumn(0, delegate)

        # print(['1','1','2','2','3','3','4','4']*6)
        self.table.setVerticalHeaderLabels(['1','1','2','2','3','3','4','4']*6)
        # self.table.setSpan(0, 0, 8, 1)
        # QTableWidgetItem(lesson).data()
        days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
        for num_day, day in enumerate(days):
            self.table.setSpan(num_day*8, 0, 8, 1)
            item = QTableWidgetItem(day)
            item.setTextAlignment(QtCore.Qt.AlignVCenter | QtCore.Qt.AlignHCenter)
            self.table.setItem(num_day*8, 0, item)

        shift_group = 1
        if TimeTable != {}:
            for num_group in range(len(groups)):
                for week in TimeTable[groups[num_group]]:
                    for day in TimeTable[groups[num_group]][week]:
                        for num_lesson in TimeTable[groups[num_group]][week][day]:
                            lessonID = TimeTable[groups[num_group]][week][day][num_lesson]
                            lesson, teacher, audienceNumber = LessonID_teacher[lessonID]
                            data_for_cell = lesson + '\n' + teacher
                            try:
                                # if week == 1 and self.table.item(day*2*4 + num_lesson*2, num_group+shift_group).text() == data_for_cell:   # пары верхней и нижней недели совпадают
                                #     self.table.setSpan(day*2*4 + num_lesson*2, num_group+shift_group, 2, 1)
                                # else:
                                self.table.setItem(day*2*4 + num_lesson*2 + week, num_group+shift_group, QTableWidgetItem(data_for_cell))
                            except Exception:
                                self.table.setItem(day*2*4 + num_lesson*2 + week, num_group+shift_group, QTableWidgetItem(data_for_cell))

        for col in range(shift_group, self.table.columnCount()):
            for row in range(0, self.table.rowCount(), 2):
                if (self.table.item(row, col) == None and self.table.item(row+1, col) == None) or (self.table.item(row, col) != None and self.table.item(row+1, col) != None and self.table.item(row,col).text() == self.table.item(row+1,col).text()):
                    self.table.setSpan(row, col, 2, 1)

        # self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)


class WorkingTime(QDialog):
    """Расписание учителя"""
    def __init__(self, name, id, DB):
        super().__init__()
        self.DB = DB
        self.attendance = self.DB.getTeacherTimeWork(id)
        self.id_teacher = id
        self.setWindowTitle('Расписание учителя')
        self.setWindowIcon(QIcon("img/teacher.png"))
        self.setWindowFlags(Qt.WindowFullscreenButtonHint)

        self.resize(900, 500)
        lay = QVBoxLayout()
        self.setLayout(lay)
        text = QLabel()
        text.setText('Расписание учителя <<' + name + '>>')
        text.setAlignment(QtCore.Qt.AlignCenter)
        lay.addWidget(text)
        self.table = QTableWidget(8, 6)
        # self.table.setColumnCount(6)
        # self.table.setRowCount(8)
        lay.addWidget(self.table)
        # self.table.setSpan(2, 2, 2, 1)
        # self.table.horizontalHeader().setStretchLastSection(True)
        # self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        self.table.cellClicked.connect(lambda row, column: self.click_work_time(row, column))
        self.table.setHorizontalHeaderLabels(['Понедельник', "Вторник", "Среда", "Четверг", "Пятница", "Суббота"])
        self.table.setVerticalHeaderLabels(['1', '1', '2', '2', '3', '3', '4', '4'])
        self.table.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.red = [255, 0, 0]
        self.dark_red = [150, 0, 0]
        self.green = [0, 255, 0]
        self.dark_green = [0, 179, 0]
        self.paint_table_teacher_attendance()
        self.insert_data_in_table()

        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        for row in range(6):
            self.table.horizontalHeader().setSectionResizeMode(row, 1)
        for col in range(8):
            self.table.verticalHeader().setSectionResizeMode(col, 1)

        self.table.setMouseTracking(True)
        self.current_hover = [0, 0]
        self.table.cellEntered.connect(self.cellHover)
        self.table.setSelectionMode(False)
        # self.table.clear()  # убирает содержимое ячеек, но объединенные оставляет
        # self.table.setColumnCount(0)
        # self.table.setRowCount(0)
        # self.table.setColumnCount(6)
        # self.table.setRowCount(8)
        # for row in range(8):
        #     for col in range(6):
        #         print(self.table.item(row, col))

    def cellHover(self, row, column):   # при наведении
        old_item = self.table.item(self.current_hover[0], self.current_hover[1])
        if self.current_hover != [row, column]:
            day, lesson = old_item.column(), old_item.row() // 2
            if self.attendance[day][lesson]:  # работает --> зеленый
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.green))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.green))
            else:  # не работает --> красный
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.red))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.red))

            day, lesson = column, row // 2
            if self.attendance[day][lesson]:  # работает --> зеленый
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.dark_green))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.dark_green))
            else:  # не работает --> красный
                self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.dark_red))
                self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.dark_red))
        self.current_hover = [row, column]

    def paint_table_teacher_attendance(self, set=True):    # закрасить при открытии окна
        attendance = self.attendance
        for day in range(6):
            for lesson in range(4):
                if set:
                    self.table.setItem(lesson * 2, day, QTableWidgetItem())
                    self.table.setItem(lesson * 2 + 1, day, QTableWidgetItem())
                if attendance[day][lesson]:
                    self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.green))
                    self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.green))
                else:
                    self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.red))
                    self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.red))

    def click_work_time(self, row, column):
        attendance = self.attendance
        day, lesson = column, row // 2
        attendance[day][lesson] = -1 if attendance[day][lesson] != -1 else 0

        if attendance[day][lesson]:  # работает --> зеленый
            self.DB.add_work_day(day, lesson, self.id_teacher)    # добавляем в бд
            setStyle(self, extraStyle="""QTreeView,QListView,QTableView{
                border: 3px solid #B8B8B8;
            }
            """, selectCell=False)
            self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.green))
            self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.green))
        else:  # не работает --> красный
            self.DB.delete_work_day(self.id_teacher, day, lesson)   # добавляем в бд
            setStyle(self, extraStyle="""QTreeView,QListView,QTableView{
                            border: 3px solid #B8B8B8;
                            }""", selectCell=False)
            self.table.item(lesson * 2, day).setBackground(QtGui.QColor(*self.red))
            self.table.item(lesson * 2 + 1, day).setBackground(QtGui.QColor(*self.red))

    def insert_data_in_table(self):
        """Вставляем расписание учителя"""
        teacherTimeWork = self.DB.getTeacherTimeWorkLesson(self.id_teacher)
        # print(teacherTimeWork)
        for week in teacherTimeWork:
            for day in teacherTimeWork[week]:
                for num_lesson in teacherTimeWork[week][day]:
                    lesson, groups = teacherTimeWork[week][day][num_lesson]
                    data_for_set = lesson+'\n'+', '.join(groups)
                    row = num_lesson*2+week
                    col = day
                    self.table.item(row, col).setText(data_for_set)
        # если одинаковые ячейки, то оъединяем их
        for col in range(self.table.columnCount()):
            for row in range(0, self.table.rowCount(), 2):
                if self.table.item(row,col).text() == self.table.item(row+1,col).text():
                    self.table.setSpan(row, col, 2, 1)


def setStyle(ui, extraStyle='', selectCell=True):
    """Стиль не для всей программы"""
    style = """
    QPushButton { text-align: center;}
        QLabel {background-color: rgb(255,128,0);}
        QWidget {font-size:16px; font-family:Comic Sans MS;
    }

    QTableView::item:selected, QListView::item:selected, QTreeView::item:selected {
        color: #F0F0F0;
        background: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0 #5d5e61, stop:1 #adadad);
    }

    QTreeView,QListView,QTableView{
        border: 1px solid #B8B8B8;
        selection-background-color: #454648;
        selection-color: #F0F0F0;
    }
    """
    if selectCell:
        style += """QTableView::item:hover, QListView::item:hover, QTreeView::item:hover {
        color: #F0F0F0;
        background: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0 #B8B8B8, stop:1 #D6D6D6);
        }"""
    ui.setStyleSheet(style+extraStyle)


class test():
    def __init__(self, num):
        self.num = num


if __name__ == '__main__':
    tt = {'ИСП11-23': {0: {0: {1: 11, 3: 14}, 1: {1: 14, 2: 11}, 2: {0: 14, 1: 14, 2: 14}, 3: {0: 11, 1: 11, 2: 14, 3: 14}, 4: {0: 14, 1: 14, 2: 11}, 5: {0: 11, 1: 11, 2: 11, 3: 11}}, 1: {0: {0: 11, 3: 11}, 1: {0: 14, 1: 11, 2: 11, 3: 11}, 2: {0: 11, 1: 11, 2: 11, 3: 11}, 3: {1: 14, 2: 14}, 4: {1: 14, 2: 14}, 5: {1: 14, 2: 11, 3: 11}}}, 'ИСП12-23': {0: {0: {3: 14, 0: 12, 1: 12, 2: 12}, 1: {1: 14, 0: 12}, 2: {0: 14, 1: 14, 2: 14}, 3: {2: 12, 3: 14, 0: 12, 1: 12}, 4: {0: 14, 1: 14, 2: 12}, 5: {0: 12, 1: 12}}, 1: {1: {0: 14, 1: 12}, 3: {1: 14, 2: 12, 0: 12}, 4: {1: 14, 2: 14, 0: 12, 3: 12}, 0: {0: 12, 1: 12, 2: 12}, 2: {0: 12, 1: 12}, 5: {1: 14, 2: 12, 3: 12}}}}
    new_tt = {}
    for group in tt:
        new_tt[group] = [None]*48
        for week in tt[group]:
            for day in tt[group][week]:
                for num_les in tt[group][week][day]:
                    new_tt[group][day*4 + week*24 + num_les] = tt[group][week][day][num_les]
    print(new_tt)
    print(round(0.11, 3))
    # values = [test(333), test(2), test(0), test(100)]
    # print(values.__getitem__(-1))
    # print(max(range(len(values)), key = lambda i: values[i].num))
    # print(random.sample(range(4), 2))
    # a = [0,0,0,0,1,2,3,4,0,0,0,0]
    # print(a)
    # num_day = 1
    # day = a[num_day*4:num_day*8]
    # print(day)
    # random.shuffle(day)
    # a[num_day*4:num_day*8] = day
    # print(a)
    # print([1,3,2].sort())
    app = QtWidgets.QApplication([])
    f = open('style_gray.qss')
    style = f.read()
    app.setStyleSheet(style)
    # app.setStyleSheet(PyQt5_stylesheets.load_stylesheet_pyqt5(style="style_gray"))
    f.close()

    win = MainWindow()
    win.setWindowTitle('Расписание')


    sys.exit(app.exec_())